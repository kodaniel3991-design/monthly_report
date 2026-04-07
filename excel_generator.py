"""
월차보고 시스템 - Excel 보고서 생성기
기존 파일과 동일한 양식으로 출력
openpyxl 사용 / 단위: 천원
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Optional


# ── 스타일 상수 ───────────────────────────────────────────────────────────
FONT_NAME = "Arial"

def _font(bold=False, size=9, color="000000"):
    return Font(name=FONT_NAME, bold=bold, size=size, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

NUM_FMT  = '#,##0'           # 정수 천 단위
PCT_FMT  = '0.00%'           # 퍼센트
ZERO_FMT = '#,##0;(#,##0);-' # 0은 '-' 표시

HDR_FILL  = _fill("D3D1C7")  # 헤더 회색
SUBHDR    = _fill("F1EFE8")  # 서브헤더 연회색
RKM_FILL  = _fill("E6F1FB")  # RKM 파랑 계열
HKMC_FILL = _fill("E1F5EE")  # HKMC 초록 계열
TOTAL_FILL= _fill("FAEEDA")  # 합계 황색


def _w(ws, col_idx: int, width: float):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def _cell(ws, row, col, value=None, bold=False, fill=None,
          align_h="center", fmt=None, border=True, size=9):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold, size=size)
    if fill:
        c.fill = fill
    c.alignment = _align(align_h)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = _border()
    return c


def _merge(ws, r1, c1, r2, c2, value=None, bold=False, fill=None,
           align_h="center", size=9):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2,   end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    c.font = _font(bold=bold, size=size)
    if fill:
        c.fill = fill
    c.alignment = _align(align_h)
    return c


# ── 손익실적 시트 ─────────────────────────────────────────────────────────

def _build_pl_sheet(wb: Workbook, year: int, month: int,
                    pl_data: dict, plan_data: dict = None):
    """
    (3). X月 損益計劃 對 實績 (總括) 시트 생성
    컬럼 구성:
      A   : 구분 (행 라벨)
      B-C : 사업계획 (금액, %)
      D-E : 수정계획 (금액, %)
      F-G : 김해공장 실적 (금액, %)
      H-I : 부산공장 실적
      J-K : RK 합계
      L-M : 울산공장 실적
      N-O : 김해2공장 실적
      P-Q : HKMC 합계
      R-S : 전체합계
    """
    ws = wb.create_sheet(f"{month}月 損益實績")
    ws.sheet_view.zoomScale = 85

    # 열 너비
    widths = [22, 12,7, 12,7, 12,7, 12,7, 13,7, 12,7, 12,7, 13,7, 14,7]
    for i, w in enumerate(widths, 1):
        _w(ws, i, w)

    # ── 타이틀 ────────────────────────────────────────────────────────────
    _merge(ws, 1,1,1,19,
           f"({month}). {month}月 損益計劃 對 實績  (總括)",
           bold=True, size=11)
    ws.row_dimensions[1].height = 18

    _merge(ws, 2,1,2,15, "진양오토모티브 김해", bold=True)
    _merge(ws, 2,16,2,19, f"(單位 : 阡Won, 臺, %)  {year}年 {month}月", size=8)

    # ── 컬럼 헤더 (row 3~4) ───────────────────────────────────────────────
    headers = [
        (3,1,4,1, "區  分",   HDR_FILL),
        (3,2,3,3, "事業計劃", HDR_FILL),
        (3,4,3,5, "修正計劃", HDR_FILL),
        (3,6,3,7, "김해공장", RKM_FILL),
        (3,8,3,9, "부산공장", RKM_FILL),
        (3,10,3,11,"RK 합계", RKM_FILL),
        (3,12,3,13,"울산공장",HKMC_FILL),
        (3,14,3,15,"김해2공장",HKMC_FILL),
        (3,16,3,17,"HKMC 합계",HKMC_FILL),
        (3,18,3,19,"전  체  합  계",TOTAL_FILL),
    ]
    for r1,c1,r2,c2,v,f in headers:
        _merge(ws,r1,c1,r2,c2, v, bold=True, fill=f)

    sub_labels = ["금  액","%(賣出)"] * 9
    for i, lab in enumerate(sub_labels):
        _cell(ws, 4, 2+i, lab, fill=HDR_FILL, bold=True, size=8)

    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 14

    # ── 데이터 행 정의 ─────────────────────────────────────────────────────
    def get_pl(key, factory, default=0):
        return pl_data.get(f"{key}_{factory}", default) or default

    def sales(f):
        return get_pl("sales_prod",f) + get_pl("sales_out",f)

    def var_cost(f):
        return (get_pl("inv_diff",f) + get_pl("material",f) +
                get_pl("mfg_welfare",f) + get_pl("mfg_power",f) +
                get_pl("mfg_trans",f) + get_pl("mfg_repair",f) +
                get_pl("mfg_supplies",f) + get_pl("mfg_fee",f) +
                get_pl("mfg_other",f) + get_pl("selling_trans",f) +
                get_pl("merch_purchase",f))

    def fixed_cost(f):
        return (get_pl("labor_salary",f) + get_pl("labor_wage",f) +
                get_pl("labor_bonus",f) + get_pl("labor_retire",f) +
                get_pl("labor_outsrc",f) + get_pl("staff_salary",f) +
                get_pl("staff_bonus",f) + get_pl("staff_retire",f) +
                get_pl("fix_depr",f) + get_pl("fix_lease",f) +
                get_pl("fix_outsrc",f) + get_pl("fix_other",f))

    def va(f):
        s = sales(f); vc = var_cost(f); mw = get_pl("mfg_welfare",f)
        return s - vc + mw

    # 공장 코드 순서: 김해, 부산, RK합, 울산, 김해2, HKMC합, 전체
    FACS = ["gimhae","busan","ulsan","gimhae2"]
    def rkm(getter):  return getter("gimhae") + getter("busan")
    def hkmc(getter): return getter("ulsan")  + getter("gimhae2")
    def total(getter):return getter("gimhae") + getter("busan") + getter("ulsan") + getter("gimhae2")

    def vals_7(getter):
        """김해, 부산, RK합, 울산, 김해2, HKMC합, 전체 순"""
        return [getter("gimhae"), getter("busan"), rkm(getter),
                getter("ulsan"),  getter("gimhae2"), hkmc(getter), total(getter)]

    # plan_data 에서 계획치 (없으면 0)
    def plan(key):
        return (plan_data or {}).get(key, 0) or 0

    rows = [
        # (라벨1, 라벨2, 라벨3, getter_or_values, is_header)
        ("販賣數量(臺)", None, None, lambda f: get_pl("qty",f), False),
        ("生産金額",     None, None, lambda f: get_pl("prod",f), False),
        ("賣出額",       None, None, sales, False),
        (None, "生産品",  None, lambda f: get_pl("sales_prod",f), False),
        (None, "外注品",  None, lambda f: get_pl("sales_out",f), False),
        ("變動費",       None, None, var_cost, True),
        (None, "製品在庫增減差", None, lambda f: get_pl("inv_diff",f), False),
        (None, "材料費",  None, lambda f: get_pl("material",f), False),
        (None, "製造經費",None, lambda f: (get_pl("mfg_welfare",f)+get_pl("mfg_power",f)+
                                          get_pl("mfg_trans",f)+get_pl("mfg_repair",f)+
                                          get_pl("mfg_supplies",f)+get_pl("mfg_fee",f)+
                                          get_pl("mfg_other",f)), False),
        (None, None, "複利厚生費", lambda f: get_pl("mfg_welfare",f), False),
        (None, None, "電力費",     lambda f: get_pl("mfg_power",f),   False),
        (None, None, "運搬費",     lambda f: get_pl("mfg_trans",f),   False),
        (None, None, "修繕費",     lambda f: get_pl("mfg_repair",f),  False),
        (None, None, "消耗品費",   lambda f: get_pl("mfg_supplies",f),False),
        (None, None, "支給手數料", lambda f: get_pl("mfg_fee",f),     False),
        (None, None, "其他",       lambda f: get_pl("mfg_other",f),   False),
        (None, "販賣運搬費",None,  lambda f: get_pl("selling_trans",f),False),
        (None, "商品買入", None,   lambda f: get_pl("merch_purchase",f),False),
        ("限界利益",     None, None, lambda f: sales(f)-var_cost(f), True),
        ("固定費",       None, None, fixed_cost, True),
        (None, "勞務費", None, lambda f: (get_pl("labor_salary",f)+get_pl("labor_wage",f)+
                                          get_pl("labor_bonus",f)+get_pl("labor_retire",f)+
                                          get_pl("labor_outsrc",f)), False),
        (None, None, "給料",       lambda f: get_pl("labor_salary",f), False),
        (None, None, "賃金",       lambda f: get_pl("labor_wage",f),   False),
        (None, None, "賞與金",     lambda f: get_pl("labor_bonus",f),  False),
        (None, None, "退充轉入額", lambda f: get_pl("labor_retire",f), False),
        (None, None, "外主用役費", lambda f: get_pl("labor_outsrc",f), False),
        (None, "人件費",None, lambda f: (get_pl("staff_salary",f)+get_pl("staff_bonus",f)+
                                         get_pl("staff_retire",f)), False),
        (None, None, "給料",       lambda f: get_pl("staff_salary",f), False),
        (None, None, "賞與金",     lambda f: get_pl("staff_bonus",f),  False),
        (None, None, "退充轉入額", lambda f: get_pl("staff_retire",f), False),
        (None, "製造·管理經費",None,lambda f: (get_pl("fix_depr",f)+get_pl("fix_lease",f)+
                                              get_pl("fix_outsrc",f)+get_pl("fix_other",f)), False),
        (None, None, "減價償却費", lambda f: get_pl("fix_depr",f),    False),
        (None, None, "支給賃借料", lambda f: get_pl("fix_lease",f),   False),
        (None, None, "外主加工費", lambda f: get_pl("fix_outsrc",f),  False),
        (None, None, "其他 經費",  lambda f: get_pl("fix_other",f),   False),
        ("營業利益",    None, None, lambda f: sales(f)-var_cost(f)-fixed_cost(f), True),
        ("附加價値",    None, None, va, True),
    ]

    row_idx = 5
    for (l1, l2, l3, getter, is_hdr) in rows:
        ws.row_dimensions[row_idx].height = 15

        # 라벨
        fill_row = HDR_FILL if is_hdr else None
        indent = 0 if l1 else (4 if l2 else 8)
        label  = l1 or l2 or l3 or ""
        _merge(ws, row_idx,1,row_idx,1, label,
               bold=is_hdr, fill=fill_row, align_h="left")

        # 계획치 (사업계획, 수정계획) - 미구현시 빈칸
        _cell(ws, row_idx, 2, None, fmt=ZERO_FMT)
        _cell(ws, row_idx, 3, None, fmt=PCT_FMT)
        _cell(ws, row_idx, 4, None, fmt=ZERO_FMT)
        _cell(ws, row_idx, 5, None, fmt=PCT_FMT)

        # 실적: 김해(6-7), 부산(8-9), RK합(10-11), 울산(12-13), 김해2(14-15), HKMC합(16-17), 전체(18-19)
        col_pairs = [(6,7),(8,9),(10,11),(12,13),(14,15),(16,17),(18,19)]
        fills_7   = [RKM_FILL, RKM_FILL, RKM_FILL,
                     HKMC_FILL, HKMC_FILL, HKMC_FILL, TOTAL_FILL]

        computed = vals_7(getter)
        totals_for_pct = [sales("gimhae"), sales("busan"), rkm(sales),
                          sales("ulsan"), sales("gimhae2"), hkmc(sales), total(sales)]

        for i, ((ca, cb), f_fill) in enumerate(zip(col_pairs, fills_7)):
            v = computed[i]
            s_ref = totals_for_pct[i]
            pct_val = v / s_ref if s_ref and l1 != "販賣數量(臺)" else None
            _cell(ws, row_idx, ca, v,
                  fill=f_fill if is_hdr else None, fmt=ZERO_FMT)
            _cell(ws, row_idx, cb, pct_val,
                  fill=f_fill if is_hdr else None,
                  fmt=PCT_FMT if pct_val is not None else "@")

        row_idx += 1

    # 영업외 섹션
    ws.row_dimensions[row_idx].height = 15
    _cell(ws, row_idx, 1, "營業外收益")
    _cell(ws, row_idx, 18, pl_data.get("non_op_income", 0), fmt=ZERO_FMT)
    row_idx += 1
    _cell(ws, row_idx, 1, "  利子收益")
    _cell(ws, row_idx, 18, pl_data.get("interest_income", 0), fmt=ZERO_FMT)
    row_idx += 1
    _cell(ws, row_idx, 1, "營業外費用")
    _cell(ws, row_idx, 18, pl_data.get("non_op_expense", 0), fmt=ZERO_FMT)
    row_idx += 1
    _cell(ws, row_idx, 1, "  利子費用")
    _cell(ws, row_idx, 18, pl_data.get("interest_expense", 0), fmt=ZERO_FMT)
    row_idx += 1

    non_op = (pl_data.get("non_op_income",0) or 0) - (pl_data.get("non_op_expense",0) or 0)
    _cell(ws, row_idx, 1,  "經常利益", bold=True, fill=HDR_FILL)
    _cell(ws, row_idx, 18, (total(lambda f: sales(f)-var_cost(f)-fixed_cost(f)) + non_op),
          bold=True, fill=HDR_FILL, fmt=ZERO_FMT)

    ws.freeze_panes = "B5"
    return ws


# ── 노동생산성 시트 ───────────────────────────────────────────────────────

def _build_labor_sheet(wb: Workbook, year: int, month: int,
                       lp_total, lp_rkm, lp_hkmc, labor_input):
    ws = wb.create_sheet(f"{month}月 勞動生産性")
    ws.sheet_view.zoomScale = 85

    widths = [20, 22, 14, 9, 14, 9, 14, 9, 14, 9]
    for i, w in enumerate(widths, 1):
        _w(ws, i, w)

    _merge(ws,1,1,1,10, f"({month}). {month}月 勞動生産性 實績", bold=True, size=11)
    _merge(ws,2,1,2,7, "진양오토모티브 김해")
    _merge(ws,2,8,2,10, f"(單位 : 阡Won, 名, %)  {year}年 {month}月", size=8)

    # 헤더
    hdrs = [
        (3,1,4,2,"區  分",HDR_FILL), (3,3,3,4,"事業計劃",HDR_FILL),
        (3,5,3,6,"修正計劃",HDR_FILL),(3,7,3,8,"實績",HDR_FILL),
        (3,9,3,10,"累計 實績",HDR_FILL),
    ]
    for r1,c1,r2,c2,v,f in hdrs:
        _merge(ws,r1,c1,r2,c2,v,bold=True,fill=f)
    for col in [3,5,7,9]:
        _cell(ws,4,col,"資料",fill=HDR_FILL,bold=True,size=8)
        _cell(ws,4,col+1,"指標",fill=HDR_FILL,bold=True,size=8)
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 14

    def row(ws, r, label, formula, val_data, val_idx, idx_data, idx_idx,
            val_fmt=ZERO_FMT, idx_fmt=PCT_FMT):
        ws.row_dimensions[r].height = 15
        _cell(ws,r,1,label, align_h="left")
        _cell(ws,r,2,formula, size=8, align_h="left")
        _cell(ws,r,3,None,fmt=val_fmt)   # 사업계획 자료
        _cell(ws,r,4,None,fmt=idx_fmt)   # 사업계획 지표
        _cell(ws,r,5,None,fmt=val_fmt)   # 수정계획 자료
        _cell(ws,r,6,None,fmt=idx_fmt)   # 수정계획 지표
        _cell(ws,r,7,val_data,fmt=val_fmt,fill=SUBHDR)   # 실적 자료
        _cell(ws,r,8,idx_data,fmt=idx_fmt,fill=SUBHDR)   # 실적 지표
        _cell(ws,r,9,None,fmt=val_fmt)   # 누계 자료
        _cell(ws,r,10,None,fmt=idx_fmt)  # 누계 지표

    r = 5
    lp = lp_total
    row(ws,r,"附加價値率","附加價値額 ÷ 賣出額",
        lp.value_added, None, lp.value_added_ratio, None,
        ZERO_FMT, PCT_FMT); r+=1
    row(ws,r,None,"÷ 賣出額",
        lp.sales, None, None, None); r+=1

    row(ws,r,"勞動生産性","附加價値額 ÷ 常時從業員數",
        lp.value_added, None, lp.labor_productivity, None,
        ZERO_FMT, ZERO_FMT); r+=1
    row(ws,r,None,"÷ 常時從業員數",
        lp.employees, None, None, None); r+=1

    row(ws,r,"勤勞所得 配分率","勞務費,給與,賞與",
        lp.labor_cost, None, lp.labor_income_ratio, None,
        ZERO_FMT, PCT_FMT); r+=1
    row(ws,r,None,"÷ 附加價値額",
        lp.value_added, None, None, None); r+=1
    row(ws,r,None,"退職,福利費",
        lp.retire_cost, None, lp.retire_ratio, None); r+=1
    row(ws,r,None,"÷ 附加價値額",
        lp.value_added, None, None, None); r+=1
    row(ws,r,None,"人件費",
        lp.labor_cost+lp.retire_cost, None, lp.total_personnel_ratio, None); r+=1
    row(ws,r,None,"÷ 附加價値額",
        lp.value_added, None, None, None); r+=1

    row(ws,r,"賣出額對比 人件費率","勞務費,給與,賞與",
        lp.labor_cost, None, lp.labor_cost_to_sales, None,
        ZERO_FMT, PCT_FMT); r+=1
    row(ws,r,None,"÷ 賣出額",
        lp.sales, None, None, None); r+=1

    row(ws,r,"1人當 賃金水準","勞務費,給與,賞與(A)",
        lp.labor_cost, None, lp.wage_per_person, None,
        ZERO_FMT, ZERO_FMT); r+=1
    row(ws,r,None,"÷ 常時從業員數",
        lp.employees, None, None, None); r+=1
    row(ws,r,None,"退職金(B)",
        lp.retire_prod, None, lp.retire_per_person, None); r+=1
    row(ws,r,None,"÷ 常時從業員數",
        lp.employees, None, None, None); r+=1

    ws.freeze_panes = "C5"
    return ws


# ── 업계동향 시트 ─────────────────────────────────────────────────────────

def _build_news_sheet(wb: Workbook, year: int, month: int,
                      news_items: list, top_models: list, market_share: dict):
    ws = wb.create_sheet(f"{month}月 업계동향")
    ws.sheet_view.zoomScale = 90

    _w(ws, 1, 6)
    _w(ws, 2, 6)
    _w(ws, 3, 30)
    _w(ws, 4, 70)

    r = 1
    _merge(ws,r,1,r,4, "▣. 자동차산업 업계 동향", bold=True, size=12)
    ws.row_dimensions[r].height = 20
    r += 2

    # 회사별 뉴스
    companies_order = ["르노코리아","GM Korea","현대자동차","업계이슈"]
    comp_num = {c: i+1 for i, c in enumerate(companies_order)}

    from itertools import groupby
    by_company = {}
    for item in news_items:
        by_company.setdefault(item["company"], []).append(item)

    for company in companies_order:
        items = by_company.get(company, [])
        if not items:
            continue
        num = comp_num.get(company, "")
        ws.row_dimensions[r].height = 15
        _merge(ws,r,2,r,2, str(num), bold=True)
        _merge(ws,r,3,r,4, company, bold=True, fill=SUBHDR, align_h="left")
        r += 1

        for item in sorted(items, key=lambda x: x["seq"]):
            ws.row_dimensions[r].height = 15
            _merge(ws,r,3,r,4, item["headline"], align_h="left")
            r += 1
            if item.get("source"):
                _cell(ws,r,4, f"<{item['source']}>", align_h="right", size=8)
                r += 1
            if item.get("content"):
                content_cell = ws.cell(row=r, column=4, value=item["content"])
                content_cell.alignment = _align("left", wrap=True)
                content_cell.font = _font(size=9)
                ws.row_dimensions[r].height = max(60, len(item["content"]) // 2)
                r += 1
            r += 1
        r += 1

    # 판매 TOP10
    if top_models:
        _merge(ws,r,2,r,4, "5. 국내 자동차 판매현황 및 시장점유율", bold=True, fill=SUBHDR)
        r += 1
        headers = ["순위","모델명","소속사","판매(대)"]
        fills_h = [HDR_FILL]*4
        for i, (h, f) in enumerate(zip(headers, fills_h)):
            _cell(ws,r,i+1, h, bold=True, fill=f)
        r += 1
        for t in sorted(top_models, key=lambda x: x["rank"]):
            _cell(ws,r,1, t["rank"])
            _cell(ws,r,2, t["model"], align_h="left")
            _cell(ws,r,3, t["company"], align_h="left")
            _cell(ws,r,4, t["qty"], fmt=NUM_FMT)
            r += 1
        r += 2

    # 시장점유율
    if market_share:
        _merge(ws,r,2,r,4, "시장점유율 (%)", bold=True, fill=SUBHDR)
        r += 1
        for comp, pct in market_share.items():
            _cell(ws,r,2, comp, align_h="left")
            _cell(ws,r,3, f"{pct:.1f}%")
            r += 1

    return ws


# ── 노동생산성 템플릿 기반 출력 ───────────────────────────────────────────

def fill_labor_productivity_template(
    template_path: str,
    pl_data: dict,
    labor_data: dict,
    labor_input,
    year: int = 2026,
    month: int = 3,
) -> bytes:
    """
    기존 노동생산성 Excel 템플릿을 열어 계산값을 채워넣는다.
    - 사업부별 시트: C3(RKM), C5(HKMC), C7(합계) + 비율(C4,C6,C8)
    - 총괄 시트: C7(실적), C8(지표)
    - 월별 컬럼(C14~C25): 1~N월 누계용 데이터
    """
    import openpyxl

    # 원본을 data_only=True로 읽어서 수식의 계산 결과값을 보존
    wb_vals = openpyxl.load_workbook(template_path, data_only=True)
    cached_values = {}  # {(sheet_name, row, col): value}
    for sn in wb_vals.sheetnames:
        ws_v = wb_vals[sn]
        for row in ws_v.iter_rows(min_row=1, max_row=ws_v.max_row,
                                   min_col=1, max_col=ws_v.max_column):
            for cell in row:
                if cell.value is not None:
                    cached_values[(sn, cell.row, cell.column)] = cell.value
    wb_vals.close()

    wb = openpyxl.load_workbook(template_path)
    sheets = wb.sheetnames

    # 시트 인덱스: [0]=총괄, [1]=사업부별, [2]=노무비(25-05), [3]=노무비(인원,근무시간)
    if len(sheets) < 2:
        raise ValueError(f"템플릿에 시트가 2개 이상이어야 합니다 (현재 {len(sheets)}개)")

    ws_div = wb[sheets[1]]  # 사업부별 시트
    ws_total = wb[sheets[0]]  # 총괄 시트

    # ── (0) 모든 수식을 원본 계산값으로 먼저 교체 ──
    for ws in [ws_total, ws_div]:
        sn = ws.title
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = cached_values.get((sn, cell.row, cell.column))

    # ── 계산 ──
    from calculator import (
        build_factory_pl_from_db, build_labor_input_from_db,
        calc_value_added, _sum_factories
    )

    gimhae  = build_factory_pl_from_db(pl_data, "gimhae")
    busan   = build_factory_pl_from_db(pl_data, "busan")
    ulsan   = build_factory_pl_from_db(pl_data, "ulsan")
    gimhae2 = build_factory_pl_from_db(pl_data, "gimhae2")

    rkm  = _sum_factories("RKM",  gimhae, busan)
    hkmc = _sum_factories("HKMC", ulsan, gimhae2)

    def salary_bonus(factories):
        """급여+상여 (퇴직 제외) — 손익실적 노무비+인건비 중 급료·임금·상여·외주"""
        total = 0
        for f in factories:
            total += sum([
                pl_data.get(f"labor_salary_{f}", 0) or 0,
                pl_data.get(f"labor_wage_{f}", 0) or 0,
                pl_data.get(f"labor_bonus_{f}", 0) or 0,
                pl_data.get(f"labor_outsrc_{f}", 0) or 0,
                pl_data.get(f"staff_salary_{f}", 0) or 0,
                pl_data.get(f"staff_bonus_{f}", 0) or 0,
            ])
        return total

    def retire_welfare(factories):
        """퇴직·복리비 — 손익실적의 퇴충전입액"""
        total = 0
        for f in factories:
            total += sum([
                pl_data.get(f"labor_retire_{f}", 0) or 0,
                pl_data.get(f"staff_retire_{f}", 0) or 0,
            ])
        return total

    # ── 사업부별 시트 값 쓰기 ──
    # C3 = RKM, C5 = HKMC
    rkm_facs = ["gimhae", "busan"]
    hkmc_facs = ["ulsan", "gimhae2"]

    # 생산직 노무비 (시간당임금용: 임금 + 상여금만, 원본 방식)
    def prod_labor_wage(factories):
        """생산직 임금 합계 (급료 제외, 임금만)"""
        return sum((pl_data.get(f"labor_wage_{f}", 0) or 0) for f in factories)

    # R41 = 임금(labor_wage) + 상여금(bonus_prod) — 생산직만
    if labor_input:
        labor_rkm_41 = prod_labor_wage(rkm_facs) + labor_input.bonus_prod_rkm
        labor_hkmc_41 = prod_labor_wage(hkmc_facs) + labor_input.bonus_prod_hkmc
    else:
        labor_rkm_41 = prod_labor_wage(rkm_facs)
        labor_hkmc_41 = prod_labor_wage(hkmc_facs)

    # (1) 값 셀 — R14는 덮어쓰지 않음 (원본 템플릿의 회계팀자료 기준값 유지)
    value_cells = {
        8:  (calc_value_added(rkm),         calc_value_added(hkmc)),
        9:  (rkm.sales,                     hkmc.sales),
        12: (salary_bonus(rkm_facs),        salary_bonus(hkmc_facs)),
        # R14: 원본 값 유지 (step 0에서 cached_values로 복원됨)
        26: (retire_welfare(rkm_facs),      retire_welfare(hkmc_facs)),
        43: (rkm.prod_amount,               hkmc.prod_amount),
    }

    for row, (rkm_val, hkmc_val) in value_cells.items():
        ws_div.cell(row=row, column=3).value = rkm_val
        ws_div.cell(row=row, column=5).value = hkmc_val

    # (2) 수식 셀 → 값으로 교체 (노무비 시트가 없을 때 #REF 방지)
    #     노무비 시트를 참조하던 수식을 계산된 값으로 덮어쓴다
    if labor_input:
        formula_override = {
            # R11: 종업원수 (원래 수식: ='노무비(인원,근무시간)'!$C$6)
            11: (labor_input.rkm_employees,     labor_input.hkmc_employees),
            # R39: 생산인원 (원래 수식: ='노무비(인원,근무시간)'!$C$5)
            39: (labor_input.prod_rkm,          labor_input.prod_hkmc),
            # R40: 실작업시간 (원래 수식: ='노무비(인원,근무시간)'!$C$16)
            40: (labor_input.work_hours_rkm,    labor_input.work_hours_hkmc),
            # R41: 생산직 노무비 = 임금 + 상여 (원래 수식: =$M$40+'노무비(25-05)'!$E$8)
            41: (labor_rkm_41,                  labor_hkmc_41),
            # R47: 퇴직금 생산직 (원래 수식: ='노무비(25-05)'!$E$14)
            47: (labor_input.retire_prod_rkm,   labor_input.retire_prod_hkmc),
        }

        for row, (rkm_val, hkmc_val) in formula_override.items():
            ws_div.cell(row=row, column=3).value = rkm_val
            ws_div.cell(row=row, column=5).value = hkmc_val

    # (3) 사업부별 C7(합계) 수식 → 값으로 교체
    va_total = calc_value_added(rkm) + calc_value_added(hkmc)
    total_sales = rkm.sales + hkmc.sales
    total_sb = salary_bonus(rkm_facs) + salary_bonus(hkmc_facs)
    total_rw = retire_welfare(rkm_facs) + retire_welfare(hkmc_facs)
    total_prod = rkm.prod_amount + hkmc.prod_amount

    total_labor_41 = labor_rkm_41 + labor_hkmc_41  # 생산직 노무비 합계

    # R14: 원본 cached 값 읽기 (회계팀자료 기준, 덮어쓰지 않음)
    r14_rkm = ws_div.cell(14, 3).value or retire_welfare(rkm_facs)
    r14_hkmc = ws_div.cell(14, 5).value or retire_welfare(hkmc_facs)
    r14_total = (r14_rkm if isinstance(r14_rkm, (int,float)) else 0) + \
                (r14_hkmc if isinstance(r14_hkmc, (int,float)) else 0)

    div_c7 = {
        8:  va_total,
        9:  total_sales,
        12: total_sb,
        # R14: C7도 원본 유지
        14: r14_total,
        26: total_rw,
        43: total_prod,
    }
    if labor_input:
        div_c7.update({
            11: labor_input.total_employees,
            39: labor_input.prod_employees,
            40: labor_input.total_work_hours,
            41: total_labor_41,
            47: labor_input.retire_prod_total,
        })

    for row, val in div_c7.items():
        ws_div.cell(row=row, column=7).value = val

    # (4) 사업부별 나머지 수식 행도 값으로 (C3, C5, C7 모두)
    #     R10=부가가치, R13=부가가치, R15=부가가치, R17=부가가치 (반복 참조)
    for row in [10, 13, 15, 17]:
        ws_div.cell(row=row, column=3).value = calc_value_added(rkm)
        ws_div.cell(row=row, column=5).value = calc_value_added(hkmc)
        ws_div.cell(row=row, column=7).value = va_total

    # R16=인건비(R12+R14), R22=인건비(R18+R20) — R14는 원본 값 사용
    for row in [16, 22]:
        ws_div.cell(row=row, column=3).value = salary_bonus(rkm_facs) + r14_rkm
        ws_div.cell(row=row, column=5).value = salary_bonus(hkmc_facs) + r14_hkmc
        ws_div.cell(row=row, column=7).value = total_sb + r14_total

    # R18=R12, R24=R12 (노무비 반복)
    for row in [18, 24]:
        ws_div.cell(row=row, column=3).value = salary_bonus(rkm_facs)
        ws_div.cell(row=row, column=5).value = salary_bonus(hkmc_facs)
        ws_div.cell(row=row, column=7).value = total_sb

    # R19,R21,R23 = 매출액 (반복)
    for row in [19, 21, 23]:
        ws_div.cell(row=row, column=3).value = rkm.sales
        ws_div.cell(row=row, column=5).value = hkmc.sales
        ws_div.cell(row=row, column=7).value = total_sales

    # R20 = 퇴직복리 (반복, R14 원본값)
    ws_div.cell(row=20, column=3).value = r14_rkm
    ws_div.cell(row=20, column=5).value = r14_hkmc
    ws_div.cell(row=20, column=7).value = r14_total

    # R25,R27,R29 = 종업원수 (반복)
    if labor_input:
        for row in [25, 27, 29]:
            ws_div.cell(row=row, column=3).value = labor_input.rkm_employees
            ws_div.cell(row=row, column=5).value = labor_input.hkmc_employees
            ws_div.cell(row=row, column=7).value = labor_input.total_employees

    # R28 = A+B (A=R24=급여상여, B=R26=퇴직금from PL)
    ws_div.cell(row=28, column=3).value = salary_bonus(rkm_facs) + retire_welfare(rkm_facs)
    ws_div.cell(row=28, column=5).value = salary_bonus(hkmc_facs) + retire_welfare(hkmc_facs)
    ws_div.cell(row=28, column=7).value = total_sb + total_rw  # R26 기준 (R14와 다름)

    # R42=실작업시간, R44/R46/R48=생산인원 (반복)
    if labor_input:
        ws_div.cell(row=42, column=3).value = labor_input.work_hours_rkm
        ws_div.cell(row=42, column=5).value = labor_input.work_hours_hkmc
        ws_div.cell(row=42, column=7).value = labor_input.total_work_hours
        for row in [44, 46, 48]:
            ws_div.cell(row=row, column=3).value = labor_input.prod_rkm
            ws_div.cell(row=row, column=5).value = labor_input.prod_hkmc
            ws_div.cell(row=row, column=7).value = labor_input.prod_employees

    # R45=R41(생산직 노무비), R49=R41+R47(생산직 노무비+퇴직)
    ws_div.cell(row=45, column=3).value = labor_rkm_41
    ws_div.cell(row=45, column=5).value = labor_hkmc_41
    ws_div.cell(row=45, column=7).value = total_labor_41
    if labor_input:
        ws_div.cell(row=49, column=3).value = labor_rkm_41 + labor_input.retire_prod_rkm
        ws_div.cell(row=49, column=5).value = labor_hkmc_41 + labor_input.retire_prod_hkmc
        ws_div.cell(row=49, column=7).value = total_labor_41 + labor_input.retire_prod_total

    # ── (5) 총괄 시트 — 당월/누계 값 덮어쓰기 ──
    def safe_div(a, b):
        return a / b if b else 0

    emp = labor_input.total_employees if labor_input else 0
    prod_emp = labor_input.prod_employees if labor_input else 0
    total_wh = labor_input.total_work_hours if labor_input else 0
    retire_prod = labor_input.retire_prod_total if labor_input else 0

    # Section 1: R14는 원본 값(회계팀자료 기준) 사용
    section1 = {
        8:  (va_total,                  safe_div(va_total, total_sales)),
        9:  (total_sales,               None),
        10: (va_total,                  safe_div(va_total, emp)),
        11: (emp,                       None),
        12: (total_sb,                  safe_div(total_sb, va_total)),
        13: (va_total,                  None),
        14: (r14_total,                 safe_div(r14_total, va_total)),
        15: (va_total,                  None),
        16: (total_sb + r14_total,      safe_div(total_sb + r14_total, va_total)),
        17: (va_total,                  None),
        18: (total_sb,                  safe_div(total_sb, total_sales)),
        19: (total_sales,               None),
        20: (r14_total,                 safe_div(r14_total, total_sales)),
        21: (total_sales,               None),
        22: (total_sb + r14_total,      safe_div(total_sb + r14_total, total_sales)),
        23: (total_sales,               None),
        24: (total_sb,                  safe_div(total_sb, emp)),
        25: (emp,                       None),
        26: (total_rw,                  safe_div(total_rw, emp)),
        27: (emp,                       None),
        28: (total_sb + total_rw,       safe_div(total_sb + total_rw, emp)),
        29: (emp,                       None),
    }

    # Section 2: R41=생산직 노무비(임금+상여)
    section2 = {
        39: (prod_emp,                          None),
        40: (total_wh,                          None),
        41: (total_labor_41,                    safe_div(total_labor_41, total_wh)),
        42: (total_wh,                          None),
        43: (total_prod,                        safe_div(total_prod, prod_emp)),
        44: (prod_emp,                          None),
        45: (total_labor_41,                    safe_div(total_labor_41, prod_emp)),
        46: (prod_emp,                          None),
        47: (retire_prod,                       safe_div(retire_prod, prod_emp)),
        48: (prod_emp,                          None),
        49: (total_labor_41 + retire_prod,      safe_div(total_labor_41 + retire_prod, prod_emp)),
        50: (prod_emp,                          None),
        51: (total_prod,                        safe_div(total_prod, total_labor_41 + retire_prod)),
        52: (total_labor_41 + retire_prod,      None),
    }

    all_sections = {**section1, **section2}

    for row, (c7_val, c8_val) in all_sections.items():
        ws_total.cell(row=row, column=7).value = c7_val
        if c8_val is not None:
            ws_total.cell(row=row, column=8).value = c8_val

    # R31~R33: R32 C7 = 누계 부가가치(=$I$8), R33 C7 = 누계종업원/월수
    ws_total.cell(row=31, column=7).value = month
    # R32/R33은 누계 계산 후 아래에서 채움

    # 총괄: C9/C10 (누계) — SUBTOTAL 수식 대신 값으로
    # 누계 = 월별 합산이므로 사업부별 월별 데이터를 합산
    # 먼저 월별 데이터 수집
    monthly_vals = {}  # {row: [m1_val, m2_val, ...]}
    cum_rows = [8, 9, 11, 12, 14, 26, 39, 40, 41, 43, 47]
    for row in cum_rows:
        vals = []
        for m in range(1, month + 1):
            col = 14 + (m - 1)
            v = ws_div.cell(row=row, column=col).value
            vals.append(v if isinstance(v, (int, float)) else 0)
        monthly_vals[row] = sum(vals)

    # C9 누계 자료
    cum_c9 = {
        8:  monthly_vals.get(8, 0),
        9:  monthly_vals.get(9, 0),
        11: monthly_vals.get(11, 0),
        12: monthly_vals.get(12, 0),
        14: monthly_vals.get(14, 0),
        26: monthly_vals.get(26, 0),
        39: monthly_vals.get(39, 0),
        40: monthly_vals.get(40, 0),
        41: monthly_vals.get(41, 0),
        43: monthly_vals.get(43, 0),
        47: monthly_vals.get(47, 0),
    }
    # 파생 누계
    cum_va = cum_c9[8]
    cum_sales = cum_c9[9]
    cum_emp = cum_c9[11]
    cum_sb = cum_c9[12]
    cum_rw = cum_c9[14]
    cum_prod_emp = cum_c9[39]
    cum_wh = cum_c9[40]
    cum_labor = cum_c9[41]
    cum_prod = cum_c9[43]
    cum_retire_prod = cum_c9[47]

    cum_section1 = {
        8: cum_va, 9: cum_sales,
        10: cum_va, 11: cum_emp,
        12: cum_sb, 13: cum_va,
        14: cum_rw, 15: cum_va,
        16: cum_sb + cum_rw, 17: cum_va,
        18: cum_sb, 19: cum_sales,
        20: cum_rw, 21: cum_sales,
        22: cum_sb + cum_rw, 23: cum_sales,
        24: cum_sb, 25: cum_emp,
        26: cum_c9[26], 27: cum_emp,
        28: cum_sb + cum_c9[26], 29: cum_emp,
    }
    cum_section2 = {
        39: cum_prod_emp, 40: cum_wh,
        41: cum_labor, 42: cum_wh,
        43: cum_prod, 44: cum_prod_emp,
        45: cum_labor, 46: cum_prod_emp,
        47: cum_retire_prod, 48: cum_prod_emp,
        49: cum_labor + cum_retire_prod, 50: cum_prod_emp,
        51: cum_prod, 52: cum_labor + cum_retire_prod,
    }
    all_cum = {**cum_section1, **cum_section2}

    for row, val in all_cum.items():
        ws_total.cell(row=row, column=9).value = val

    # C10 누계 지표
    cum_c10 = {
        8:  safe_div(cum_va, cum_sales),
        10: safe_div(cum_va, safe_div(cum_emp, month)),  # 평균 종업원
        12: safe_div(cum_sb, cum_va),
        14: safe_div(cum_rw, cum_va),
        16: safe_div(cum_sb + cum_rw, cum_va),
        18: safe_div(cum_sb, cum_sales),
        20: safe_div(cum_rw, cum_sales),
        22: safe_div(cum_sb + cum_rw, cum_sales),
        24: safe_div(cum_sb, safe_div(cum_emp, month)),
        26: safe_div(cum_c9[26], safe_div(cum_emp, month)),
        28: safe_div(cum_sb + cum_c9[26], safe_div(cum_emp, month)),
        32: safe_div(cum_va, safe_div(cum_emp, month)),
        41: safe_div(cum_labor, cum_wh),
        43: safe_div(cum_prod, safe_div(cum_prod_emp, month)),
        45: safe_div(cum_labor, safe_div(cum_prod_emp, month)),
        47: safe_div(cum_retire_prod, safe_div(cum_prod_emp, month)),
        49: safe_div(cum_labor + cum_retire_prod, safe_div(cum_prod_emp, month)),
        51: safe_div(cum_prod, cum_labor + cum_retire_prod),
    }
    for row, val in cum_c10.items():
        ws_total.cell(row=row, column=10).value = val

    # R32/R33: C7=누계, C8=지표, C9=최근1년
    ws_total.cell(row=32, column=7).value = cum_va  # =$I$8 (누계 부가가치)
    ws_total.cell(row=32, column=8).value = safe_div(cum_va, safe_div(cum_emp, month))
    ws_total.cell(row=33, column=7).value = safe_div(cum_emp, month)  # 평균 종업원
    ws_total.cell(row=32, column=9).value = cum_va
    ws_total.cell(row=33, column=9).value = safe_div(cum_emp, month)

    # 총괄: C15~C26 월별 컬럼 (사업부별 참조 수식 → 값으로)
    monthly_rows = [8, 9, 11, 12, 14, 26, 32, 33, 39, 40, 41, 43, 47]
    for m in range(1, month + 1):
        src_col = 14 + (m - 1)  # 사업부별 월별 컬럼
        dst_col = 15 + (m - 1)  # 총괄 월별 컬럼 (O=15부터)
        for row in monthly_rows:
            v = ws_div.cell(row=row, column=src_col).value
            if isinstance(v, (int, float)):
                ws_total.cell(row=row, column=dst_col).value = v

    # 사업부별 C4/C6/C8 비율 채우기
    for row, (c7_val, c8_val) in all_sections.items():
        if c8_val is not None:
            # C8 = 합계 지표
            ws_div.cell(row=row, column=8).value = c8_val
            # C4 = RKM 지표, C6 = HKMC 지표
            c3 = ws_div.cell(row=row, column=3).value or 0
            c3d = ws_div.cell(row=row+1, column=3).value if row+1 <= 52 else 0
            c5 = ws_div.cell(row=row, column=5).value or 0
            c5d = ws_div.cell(row=row+1, column=5).value if row+1 <= 52 else 0
            if isinstance(c3, (int, float)) and isinstance(c3d, (int, float)) and c3d:
                ws_div.cell(row=row, column=4).value = safe_div(c3, c3d)
            if isinstance(c5, (int, float)) and isinstance(c5d, (int, float)) and c5d:
                ws_div.cell(row=row, column=6).value = safe_div(c5, c5d)

    # ── (6) 사업부별 월별 컬럼 (C14=1월 ~ C25=12월) ──
    # 누계(C9) = SUBTOTAL(109, N:Y)이므로 월별 컬럼에 값을 채워야 함
    from database import load_monthly_pl, load_monthly_labor

    def _calc_month_values(m_pl, m_lb):
        """월별 PL+Labor → 사업부별 월별 컬럼에 넣을 값 dict"""
        if not m_pl:
            return None

        g  = build_factory_pl_from_db(m_pl, "gimhae")
        b  = build_factory_pl_from_db(m_pl, "busan")
        u  = build_factory_pl_from_db(m_pl, "ulsan")
        g2 = build_factory_pl_from_db(m_pl, "gimhae2")
        r  = _sum_factories("RKM", g, b)
        h  = _sum_factories("HKMC", u, g2)

        va = calc_value_added(r) + calc_value_added(h)
        sales = r.sales + h.sales
        sb = salary_bonus(["gimhae","busan"]) if m_pl is pl_data else _salary_bonus_from(m_pl)
        rw = retire_welfare(["gimhae","busan"]) if m_pl is pl_data else _retire_welfare_from(m_pl)

        # 범용 계산 (다른 월용)
        sb_all = 0
        rw_all = 0
        lc_all = 0
        for f in ["gimhae","busan","ulsan","gimhae2"]:
            sb_all += sum([
                m_pl.get(f"labor_salary_{f}", 0) or 0,
                m_pl.get(f"labor_wage_{f}", 0) or 0,
                m_pl.get(f"labor_bonus_{f}", 0) or 0,
                m_pl.get(f"labor_outsrc_{f}", 0) or 0,
                m_pl.get(f"staff_salary_{f}", 0) or 0,
                m_pl.get(f"staff_bonus_{f}", 0) or 0,
            ])
            rw_all += sum([
                m_pl.get(f"labor_retire_{f}", 0) or 0,
                m_pl.get(f"staff_retire_{f}", 0) or 0,
            ])
            lc_all += sum([
                m_pl.get(f"labor_salary_{f}", 0) or 0,
                m_pl.get(f"labor_wage_{f}", 0) or 0,
                m_pl.get(f"labor_bonus_{f}", 0) or 0,
                m_pl.get(f"labor_retire_{f}", 0) or 0,
                m_pl.get(f"labor_outsrc_{f}", 0) or 0,
                m_pl.get(f"staff_salary_{f}", 0) or 0,
                m_pl.get(f"staff_bonus_{f}", 0) or 0,
                m_pl.get(f"staff_retire_{f}", 0) or 0,
            ])

        prod = r.prod_amount + h.prod_amount

        vals = {
            8:  va,
            9:  sales,
            12: sb_all,
            14: rw_all,
            26: rw_all,
            43: prod,
        }

        if m_lb:
            m_labor = build_labor_input_from_db(m_lb)
            vals[11] = m_labor.total_employees
            vals[39] = m_labor.prod_employees
            vals[40] = m_labor.total_work_hours
            vals[41] = lc_all
            vals[47] = m_labor.retire_prod_total

        return vals

    for m in range(1, month + 1):
        col = 14 + (m - 1)  # C14=1월, C15=2월, ...
        if m == month:
            m_pl_data = pl_data
            m_lb_data = labor_data
        else:
            m_pl_data = load_monthly_pl(year, m)
            m_lb_data = load_monthly_labor(year, m)

        if not m_pl_data:
            continue

        m_vals = _calc_month_values(m_pl_data, m_lb_data)
        if m_vals:
            for row, val in m_vals.items():
                ws_div.cell(row=row, column=col).value = val

    # R31 C7 = 월 수
    ws_div.cell(row=31, column=7).value = month
    ws_total.cell(row=31, column=7).value = month

    # ── (7) 수식 → 값 변환: 기존 값 기반으로 파생값 채우고 수식 제거 ──
    def _fill_derived_cols(ws, cols):
        """C3/C5 등 기존 값이 있는 컬럼의 수식 셀을 파생값으로 채운다.
        base rows(8,9,12,14,26,39,43)의 값을 기반으로 나머지 행 계산."""
        for col in cols:
            # 기존 값(숫자) 읽기
            va   = ws.cell(8, col).value  or 0
            sales= ws.cell(9, col).value  or 0
            sb   = ws.cell(12, col).value or 0
            rw   = ws.cell(14, col).value or 0
            rk_b = ws.cell(26, col).value or 0
            p_emp= ws.cell(39, col).value or 0
            prod = ws.cell(43, col).value or 0

            def _num(v):
                return v if isinstance(v, (int, float)) else 0
            va = _num(va); sales = _num(sales); sb = _num(sb)
            rw = _num(rw); rk_b = _num(rk_b); p_emp = _num(p_emp); prod = _num(prod)

            # 추가 행 읽기
            emp  = _num(ws.cell(11, col).value)
            wh   = _num(ws.cell(40, col).value)
            lc   = _num(ws.cell(41, col).value)
            rp   = _num(ws.cell(47, col).value)

            if va == 0 and sales == 0:
                continue  # 데이터 없는 컬럼은 건너뛰기

            # Section 1: 파생값 채우기
            derived = {
                10: va, 13: va, 15: va, 17: va,  # 부가가치 반복
                11: emp,                           # 종업원수
                16: sb + rw,                       # 인건비 = 노무비 + 퇴직
                18: sb, 24: sb,                    # 노무비 반복
                19: sales, 21: sales, 23: sales,   # 매출액 반복
                20: rw,                            # 퇴직복리 반복
                22: sb + rw,                       # 인건비 반복
                25: emp, 27: emp, 29: emp,         # 종업원수 반복
                28: sb + rk_b,                     # (A)+(B)
            }
            # Section 2: 파생값
            derived.update({
                40: wh, 42: wh,                    # 실작업시간
                41: lc,                            # 노무비
                44: p_emp, 46: p_emp, 48: p_emp, 50: p_emp,  # 생산인원
                45: lc,                            # 노무비 반복
                47: rp,                            # 퇴직금
                49: lc + rp,                       # 노무비+퇴직
                51: prod,                          # 생산금액
                52: lc + rp,                       # 노무비+퇴직 반복
            })

            for row, val in derived.items():
                cell = ws.cell(row, col)
                # 수식이거나 비어있으면 값으로 채움
                cv = cell.value
                if cv is None or (isinstance(cv, str) and cv.startswith("=")):
                    cell.value = val

            # 비율 컬럼 (col+1)
            ratio_col = col + 1
            ratios = {
                8:  safe_div(va, sales),
                10: safe_div(va, emp),
                12: safe_div(sb, va),
                14: safe_div(rw, va),
                16: safe_div(sb + rw, va),
                18: safe_div(sb, sales),
                20: safe_div(rw, sales),
                22: safe_div(sb + rw, sales),
                24: safe_div(sb, emp),
                26: safe_div(rk_b, emp),
                28: safe_div(sb + rk_b, emp),
                41: safe_div(lc, wh),
                43: safe_div(prod, p_emp),
                45: safe_div(lc, p_emp),
                47: safe_div(rp, p_emp),
                49: safe_div(lc + rp, p_emp),
                51: safe_div(prod, lc + rp),
            }
            for row, val in ratios.items():
                cell = ws.cell(row, ratio_col)
                cv = cell.value
                if cv is None or (isinstance(cv, str) and cv.startswith("=")):
                    cell.value = val

    # 총괄: C3/C4(사업계획), C5/C6(수정계획)
    _fill_derived_cols(ws_total, [3, 5])
    # 사업부별: C3/C4(RKM), C5/C6(HKMC)
    _fill_derived_cols(ws_div, [3, 5])

    # (수식은 이미 step 0에서 모두 값으로 교체됨)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 업계동향 템플릿 기반 출력 ──────────────────────────────────────────────

def fill_industry_template(
    template_path,
    year: int,
    month: int,
    news_items: list,
    top_models: list = None,
    market_share: dict = None,
) -> bytes:
    """
    업계동향 .xls 템플릿에 DB 데이터를 채워넣는다.
    xlrd + xlutils.copy로 원본 서식 유지.
    시트 구조:
      시트1: 1.르노코리아 + 2.GM Korea (각 뉴스 2건)
      시트2: 3.현대자동차 + 4.업계이슈 (각 뉴스 2건)
      시트3: 5.TOP10 판매 + 시장점유율
    """
    import xlrd
    from xlutils.copy import copy as xlcopy

    # template_path가 파일 객체(Streamlit UploadedFile)일 수 있음
    if hasattr(template_path, 'read'):
        file_contents = template_path.read()
        rb = xlrd.open_workbook(file_contents=file_contents, formatting_info=True)
    else:
        rb = xlrd.open_workbook(template_path, formatting_info=True)

    wb = xlcopy(rb)

    # ── 뉴스를 회사별로 분류 ──
    by_company = {}
    for item in news_items:
        comp = item.get("company", "")
        by_company.setdefault(comp, []).append(item)

    # 회사-시트-위치 매핑 (시트 인덱스, 위치: 1=상단, 2=하단)
    COMPANY_MAP = {
        "르노코리아": (0, 1),
        "GM Korea":   (0, 2),
        "현대자동차": (1, 1),
        "업계이슈":   (1, 2),
    }

    # 위치별 셀 매핑 (0-based row/col for xlwt)
    POS_MAP = {
        1: {  # 상단 회사
            "name_row": 3, "name_col": 3,   # R4 C4
            "headlines": [(5, 3), (13, 3)],  # R6 C4, R14 C4
            "contents":  [(7, 4, 5), (15, 4, 5)],  # (R8 C5, 5줄), (R16 C5, 5줄)
        },
        2: {  # 하단 회사
            "name_row": 22, "name_col": 3,   # R23 C4
            "headlines": [(24, 3), (31, 3)],  # R25 C4, R32 C4
            "contents":  [(26, 4, 4), (33, 4, 6)],  # (R27 C5, 4줄), (R34 C5, 6줄)
        },
    }

    def _split_content(text, max_chars=42):
        if not text:
            return []
        lines = []
        for para in text.split("\n"):
            while len(para) > max_chars:
                lines.append(para[:max_chars])
                para = para[max_chars:]
            if para.strip():
                lines.append(para)
        return lines

    # ── 시트 1, 2: 뉴스 채우기 ──
    for company, (sheet_idx, pos) in COMPANY_MAP.items():
        if sheet_idx >= rb.nsheets:
            continue
        ws = wb.get_sheet(sheet_idx)
        items = by_company.get(company, [])
        items = sorted(items, key=lambda x: x.get("seq", 1))
        layout = POS_MAP[pos]

        # 회사명
        ws.write(layout["name_row"], layout["name_col"], company)

        for news_idx, (hl_pos, ct_pos) in enumerate(
            zip(layout["headlines"], layout["contents"])
        ):
            if news_idx >= len(items):
                break
            item = items[news_idx]

            # 헤드라인 (출처 포함)
            headline = item.get("headline", "")
            source = item.get("source", "")
            if source and source not in headline:
                headline = f"{headline} <{source}>"
            ws.write(hl_pos[0], hl_pos[1], headline)

            # 내용 (여러 행으로 분할)
            content = item.get("content", "") or ""
            lines = _split_content(content)
            start_row, col, max_lines = ct_pos
            for line_idx in range(max_lines):
                if line_idx < len(lines):
                    ws.write(start_row + line_idx, col, lines[line_idx])
                else:
                    ws.write(start_row + line_idx, col, "")

    # ── 시트 3: TOP10 + 시장점유율 ──
    if rb.nsheets >= 3:
        ws3 = wb.get_sheet(2)

        # TOP10 (R9~R18: C5=순위, C6=모델명, C9=소속사, C12=판매량)
        if top_models:
            for i, model in enumerate(sorted(top_models, key=lambda x: x.get("rank", 99))):
                row = 8 + i  # 0-based
                if row > 17:
                    break
                ws3.write(row, 4, model.get("rank", i+1))
                ws3.write(row, 5, model.get("model_name", model.get("model", "")))
                ws3.write(row, 8, model.get("company", ""))
                ws3.write(row, 11, model.get("sales_qty", model.get("qty", 0)))

        # 시장점유율 (R24: C7=현대, C9=기아, C11=한국GM, C13=르노코리아, C15=KG모빌리티)
        if market_share:
            share_cols = {
                "현대": 6, "기아": 8, "GM": 10, "한국GM": 10, "한국 GM": 10,
                "르노코리아": 12, "KG모빌리티": 14, "KG": 14,
            }  # 0-based
            for comp, pct in market_share.items():
                col = share_cols.get(comp)
                if col is not None:
                    ws3.write(23, col, pct)  # R24 (0-based=23)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 공개 API ──────────────────────────────────────────────────────────────

def generate_excel(
    year: int, month: int,
    pl_data: dict,
    labor_data: dict,
    lp_total, lp_rkm, lp_hkmc,
    labor_input,
    news_items: list = None,
    top_models: list = None,
    market_share: dict = None,
    plan_data: dict = None,
) -> bytes:
    """
    모든 시트를 포함한 Excel 파일 생성 후 bytes 반환
    Streamlit download_button 에 바로 전달 가능
    """
    wb = Workbook()
    wb.remove(wb.active)   # 기본 빈 시트 제거

    _build_pl_sheet(wb, year, month, pl_data, plan_data)
    _build_labor_sheet(wb, year, month, lp_total, lp_rkm, lp_hkmc, labor_input)
    if news_items or top_models or market_share:
        _build_news_sheet(wb, year, month,
                          news_items or [], top_models or [], market_share or {})

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
