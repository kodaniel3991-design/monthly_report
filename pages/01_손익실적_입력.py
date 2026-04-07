"""
손익실적 입력 페이지
Excel 파일 업로드 → 파싱 → 미리보기 → DB 저장
단위: 천원(KRW), 대(판매수량)
"""

import streamlit as st
import pandas as pd
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from database import save_monthly_pl, load_monthly_pl
from flow_bar import render_flow_bar

st.set_page_config(page_title="손익실적 입력", layout="wide")

render_flow_bar(current_step=2)

# ── Excel 구조 정의 ──────────────────────────────────────────────────────────
# 공장별 컬럼 위치 (1-based)
FACTORY_COLS = {
    "gimhae":  8,
    "busan":   10,
    "ulsan":   14,
    "gimhae2": 16,
}

# RKM/HKMC/합계 컬럼 (미리보기용)
DIVISION_COLS = {
    "rkm":   12,
    "hkmc":  18,
    "total": 20,
}

ALL_COLS = {**FACTORY_COLS, **DIVISION_COLS}

ALL_KEYS = [
    ("gimhae",  "김해공장"),
    ("busan",   "부산공장"),
    ("rkm",     "RKM"),
    ("ulsan",   "울산공장"),
    ("gimhae2", "김해2공장"),
    ("hkmc",    "HKMC"),
    ("total",   "합계"),
]

# (DB 필드 prefix, 항목명, Excel 행번호, 유형)
# 유형: "input"=DB에 저장, "calc"=계산값(미리보기만)
PL_ROW_MAP = [
    ("qty",             "판매수량(대)",       6,  "input"),
    ("prod",            "생산금액",           7,  "input"),
    ("_sales",          "매출액",             8,  "calc"),
    ("sales_prod",      "  생산품매출",       9,  "input"),
    ("sales_out",       "  상품매출",         10, "input"),
    ("_variable_cost",  "변동비",             11, "calc"),
    ("inv_diff",        "  제품재고증감차",   12, "input"),
    ("material",        "  재료비",           13, "input"),
    ("_mfg_var_sub",    "  제조경비(변동)",   14, "calc"),
    ("mfg_welfare",     "    복리후생비★",    15, "input"),
    ("mfg_power",       "    전력비",         16, "input"),
    ("mfg_trans",       "    운반비",         17, "input"),
    ("mfg_repair",      "    수선비",         18, "input"),
    ("mfg_supplies",    "    소모품비",       19, "input"),
    ("mfg_fee",         "    지급수수료",     20, "input"),
    ("mfg_other",       "    기타(변동)",     -1, "sum_rows"),  # R21+R22+R23+R24
    ("selling_trans",   "  판매운반비",       25, "input"),
    ("merch_purchase",  "  상품매입",         26, "input"),
    ("_margin",         "한계이익",           27, "calc"),
    ("_fixed_cost",     "고정비",             28, "calc"),
    ("_labor_sub",      "  노무비",           29, "calc"),
    ("labor_salary",    "    급료",           30, "input"),
    ("labor_wage",      "    임금",           31, "input"),
    ("labor_bonus",     "    상여금",         32, "input"),
    ("labor_retire",    "    퇴충전입액",     33, "input"),
    ("labor_outsrc",    "    외주용역비",     34, "input"),
    ("_staff_sub",      "  인건비",           35, "calc"),
    ("staff_salary",    "    급료(인건비)",   36, "input"),
    ("staff_bonus",     "    상여(인건비)",   37, "input"),
    ("staff_retire",    "    퇴충(인건비)",   38, "input"),
    ("_fix_mfg_sub",    "  제조경비(고정)",   39, "calc"),
    ("fix_lease",       "    지급임차료",     41, "input"),
    ("fix_depr",        "    감가상각비",     43, "input"),
    ("fix_outsrc",      "    외주가공비",     46, "input"),
    ("fix_other",       "    기타(고정)",     -2, "sum_rows"),  # R40+R42+R44+R45+R47+R48+R49+R50
    ("_ga_sub",         "  일반관리비",       51, "calc"),
    ("_op_profit",      "영업이익",           61, "calc"),
    ("non_op_income",   "영업외수익",         62, "total_only"),
    ("interest_income", "  이자수익",         63, "total_only"),
    ("non_op_expense",  "영업외비용",         64, "total_only"),
    ("interest_expense","  이자비용",         65, "total_only"),
    ("_ordinary",       "경상이익",           66, "calc"),
]

# mfg_other에 합산할 행 (R21~R24)
MFG_OTHER_ROWS = [21, 22, 23, 24]
# fix_other에 합산할 행 (R40,R42,R44,R45,R47,R48,R49,R50)
FIX_OTHER_ROWS = [40, 42, 44, 45, 47, 48, 49, 50]

STEPS = [
    {"label": "Excel 업로드",   "num": "1"},
    {"label": "데이터 미리보기", "num": "2"},
    {"label": "저장",           "num": "3"},
    {"label": "저장 확인",      "num": "4"},
]


# ── 헬퍼 함수 ────────────────────────────────────────────────────────────────

def render_step_bar(current_step: int):
    """프로그레스 바 + 브레드크럼 칩"""
    total = len(STEPS)
    done = min(current_step + 1, total)
    pct = int(done / total * 100)

    st.markdown(
        f"<div style='background:white; border:0.5px solid #dddbd7; border-radius:8px; "
        f"padding:14px 20px 12px; box-shadow:0 1px 3px rgba(85,73,64,0.08); margin-bottom:12px;'>"
        f"<div style='display:flex; justify-content:space-between; align-items:center; margin-bottom:2px;'>"
        f"<div>"
        f"<span style='font-family:Sora,Pretendard,sans-serif; font-size:15px; font-weight:700; "
        f"color:#000; letter-spacing:-0.02em;'>손익실적 입력</span>"
        f"<span style='font-size:11px; color:#a8a9aa; margin-left:10px;'>Excel 업로드 → 미리보기 → 저장 → 확인</span>"
        f"</div>"
        f"<div style='font-family:Sora,Pretendard,sans-serif; font-size:20px; font-weight:800; "
        f"color:#554940; letter-spacing:-0.03em;'>{pct}%</div>"
        f"</div>"
        f"<div style='height:4px; background:#eceae7; border-radius:2px; margin:8px 0 10px; overflow:hidden;'>"
        f"<div style='height:100%; width:{pct}%; background:#554940; border-radius:3px; "
        f"transition:width 0.4s cubic-bezier(0.16,1,0.3,1);'></div>"
        f"</div>"
        f"<div style='display:flex; align-items:center; gap:4px; flex-wrap:wrap;'>"
        + "".join(
            f"<div style='display:inline-flex; align-items:center; gap:3px; padding:4px 10px; "
            f"border-radius:4px; border:1px solid #879a77; background:#f0f3ee; "
            f"font-size:11px; font-weight:500; color:#55654a;'>"
            f"&#10003; {s['num']}. {s['label']}</div>"
            f"<span style='color:#c5c6c7; font-size:10px;'>›</span>"
            if i < current_step else
            f"<div style='display:inline-flex; align-items:center; gap:3px; padding:4px 10px; "
            f"border-radius:4px; border:1.5px solid #554940; background:white; "
            f"font-size:11px; font-weight:700; color:#554940;'>"
            f"{s['num']}. {s['label']}</div>"
            f"<span style='color:#c5c6c7; font-size:10px;'>›</span>"
            if i == current_step else
            f"<div style='display:inline-flex; align-items:center; padding:4px 10px; "
            f"border-radius:4px; border:1px solid #dddbd7; background:#f5f4f2; "
            f"font-size:11px; color:#a8a9aa;'>"
            f"{s['num']}. {s['label']}</div>"
            + (f"<span style='color:#c5c6c7; font-size:10px;'>›</span>" if i < total - 1 else "")
            for i, s in enumerate(STEPS)
        )
        + "</div>"
        f"</div>",
        unsafe_allow_html=True
    )


def render_nav_buttons(current_step: int, max_step: int = 3):
    col_l, col_c, col_r = st.columns([1, 3, 1])
    with col_l:
        if current_step > 0:
            if st.button("< 이전", use_container_width=True, key=f"prev_{current_step}"):
                st.session_state.pl_step = current_step - 1
                st.rerun()
    with col_r:
        if current_step < max_step:
            if st.button("다음 >", type="primary", use_container_width=True, key=f"next_{current_step}"):
                st.session_state.pl_step = current_step + 1
                st.rerun()


def parse_pl_excel(uploaded_file, year: int, month: int) -> dict:
    """
    손익실적 Excel 파싱
    → {"sheet_name": ..., "data": {field_name: value, ...}, "preview": {항목: {공장: value}}}
    """
    import openpyxl

    wb = openpyxl.load_workbook(uploaded_file, data_only=True)

    # 첫 번째 시트 사용
    target_sheet = wb.sheetnames[0]
    ws = wb[target_sheet]

    def cell_val(row, col):
        v = ws.cell(row, col).value
        if v is None:
            return 0.0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0

    # ── DB에 저장할 데이터 (공장 4개 기준) ──
    db_data = {}

    for field, label, row_num, kind in PL_ROW_MAP:
        if field.startswith("_"):
            # 계산값은 DB에 저장하지 않음
            continue

        if kind == "input":
            for fcode, col in FACTORY_COLS.items():
                db_data[f"{field}_{fcode}"] = cell_val(row_num, col)

        elif kind == "sum_rows":
            # 여러 행을 합산
            if field == "mfg_other":
                sum_rows = MFG_OTHER_ROWS
            elif field == "fix_other":
                sum_rows = FIX_OTHER_ROWS
            else:
                continue
            for fcode, col in FACTORY_COLS.items():
                total = sum(cell_val(r, col) for r in sum_rows)
                db_data[f"{field}_{fcode}"] = total

        elif kind == "total_only":
            # 영업외 항목: 합계(C20) 값만 저장
            db_data[field] = cell_val(row_num, DIVISION_COLS["total"])

    # ── 미리보기용 전체 데이터 (공장+사업부+합계) ──
    preview = {}
    for field, label, row_num, kind in PL_ROW_MAP:
        row_data = {}
        if kind == "sum_rows":
            if field == "mfg_other":
                sum_rows = MFG_OTHER_ROWS
            elif field == "fix_other":
                sum_rows = FIX_OTHER_ROWS
            else:
                continue
            for key, col in ALL_COLS.items():
                row_data[key] = sum(cell_val(r, col) for r in sum_rows)
        elif kind == "total_only":
            for key, col in ALL_COLS.items():
                row_data[key] = cell_val(row_num, col)
        elif row_num > 0:
            for key, col in ALL_COLS.items():
                row_data[key] = cell_val(row_num, col)
        preview[field] = {"label": label, "values": row_data}

    wb.close()

    return {
        "sheet_name": target_sheet,
        "db_data": db_data,
        "preview": preview,
        "year": year,
        "month": month,
    }


def build_preview_df(parsed: dict) -> pd.DataFrame:
    """미리보기 데이터프레임 생성"""
    rows = []
    for field, label, row_num, kind in PL_ROW_MAP:
        pv = parsed["preview"].get(field, {})
        vals = pv.get("values", {})
        row = {"항목": label}
        for key, key_label in ALL_KEYS:
            row[key_label] = vals.get(key, 0)
        rows.append(row)
    return pd.DataFrame(rows)


def format_num_df(df: pd.DataFrame) -> pd.DataFrame:
    """숫자 컬럼을 천단위 콤마 포맷"""
    out = df.copy()
    num_cols = [c for c in out.columns if c != "항목"]
    for c in num_cols:
        out[c] = out[c].apply(lambda x: f"{x:,.0f}" if x else "0")
    return out


# ══════════════════════════════════════════════════════════════════════════════
#  메인 화면
# ══════════════════════════════════════════════════════════════════════════════

if "pl_step" not in st.session_state:
    st.session_state.pl_step = 0

step = st.session_state.pl_step

render_step_bar(step)

# ── STEP 0: Excel 업로드 ─────────────────────────────────────────────────────
if step == 0:
    col1, col2, col3 = st.columns([1, 1, 3])
    with col1:
        year = st.selectbox("연도", range(2024, 2028), index=2, key="pl_year")
    with col2:
        month = st.selectbox("월", range(1, 13), index=2, key="pl_month")

    st.session_state.pl_sel_year = year
    st.session_state.pl_sel_month = month

    st.info(
        "**지원 파일 형식**: `3월 손익 실적.xlsx` 형태의 손익실적 파일 — "
        "공장별(김해·부산·울산·김해2) 월별 실적을 자동 추출합니다."
    )

    uploaded = st.file_uploader(
        "손익실적 Excel (.xlsx)", type=["xlsx"],
        key="pl_upload", label_visibility="collapsed"
    )

    if uploaded:
        with st.spinner("Excel 파싱 중..."):
            try:
                parsed = parse_pl_excel(uploaded, year, month)
                st.session_state.pl_parsed = parsed
                st.success(f"시트 '{parsed['sheet_name']}' 에서 공장별 손익 데이터 추출 완료!")

                # 간단 요약
                db = parsed["db_data"]
                sales_g = db.get("sales_prod_gimhae", 0) + db.get("sales_out_gimhae", 0)
                sales_b = db.get("sales_prod_busan", 0) + db.get("sales_out_busan", 0)
                sales_u = db.get("sales_prod_ulsan", 0) + db.get("sales_out_ulsan", 0)
                sales_g2 = db.get("sales_prod_gimhae2", 0) + db.get("sales_out_gimhae2", 0)
                rkm_sales = sales_g + sales_b
                hkmc_sales = sales_u + sales_g2

                st.markdown(
                    f"**매출 미리보기** — "
                    f"김해: {sales_g:,.0f} / 부산: {sales_b:,.0f} / "
                    f"**RKM: {rkm_sales:,.0f}** | "
                    f"울산: {sales_u:,.0f} / 김해2: {sales_g2:,.0f} / "
                    f"**HKMC: {hkmc_sales:,.0f}**"
                )
            except Exception as e:
                st.error(f"파싱 실패: {e}")

    # 기존 저장 데이터 표시
    existing = load_monthly_pl(year, month)
    if existing and existing.get("id"):
        st.warning(f"{year}년 {month}월 기존 데이터가 있습니다. 업로드하면 덮어씁니다.")

    st.divider()
    col_l, col_c, col_r = st.columns([1, 3, 1])
    with col_r:
        can_next = "pl_parsed" in st.session_state and st.session_state.pl_parsed is not None
        if st.button("다음 >", type="primary", use_container_width=True,
                     disabled=not can_next, key="next_0"):
            st.session_state.pl_step = 1
            st.rerun()

# ── STEP 1: 미리보기 ─────────────────────────────────────────────────────────
elif step == 1:
    parsed = st.session_state.get("pl_parsed", {})
    year = parsed.get("year", st.session_state.get("pl_sel_year", 2026))
    month = parsed.get("month", st.session_state.get("pl_sel_month", 3))

    st.markdown(f"**Step 2.** {year}년 {month}월 손익실적 미리보기")
    st.caption(f"시트: {parsed.get('sheet_name', '?')} | 단위: 천원, 판매수량은 대")

    df = build_preview_df(parsed)
    st.dataframe(format_num_df(df), use_container_width=True, hide_index=True)

    # 주요 지표
    pv = parsed.get("preview", {})
    sales_vals = pv.get("_sales", {}).get("values", {})
    margin_vals = pv.get("_margin", {}).get("values", {})
    welfare_vals = pv.get("mfg_welfare", {}).get("values", {})

    st.divider()
    st.markdown("**주요 지표 검증**")
    m1, m2, m3, m4 = st.columns(4)

    rkm_sales = sales_vals.get("rkm", 0)
    hkmc_sales = sales_vals.get("hkmc", 0)
    total_sales = sales_vals.get("total", 0)
    rkm_margin = margin_vals.get("rkm", 0)

    m1.metric("RKM 매출", f"{rkm_sales:,.0f}")
    m2.metric("HKMC 매출", f"{hkmc_sales:,.0f}")
    m3.metric("합계 매출", f"{total_sales:,.0f}")
    m4.metric("RKM 한계이익", f"{rkm_margin:,.0f}")

    # 부가가치 검증
    var_cost_vals = pv.get("_variable_cost", {}).get("values", {})
    st.markdown("**부가가치 검증** (매출액 - 변동비 + 변동복리후생비)")
    v1, v2, v3 = st.columns(3)
    for col, div, label in [(v1, "rkm", "RKM"), (v2, "hkmc", "HKMC"), (v3, "total", "합계")]:
        s = sales_vals.get(div, 0)
        vc = var_cost_vals.get(div, 0)
        w = welfare_vals.get(div, 0)
        va = s - vc + w
        with col:
            st.metric(f"{label} 부가가치", f"{va:,.0f}",
                      f"{va/s*100:.1f}%" if s else "-")

    st.divider()
    render_nav_buttons(step)

# ── STEP 2: 저장 ─────────────────────────────────────────────────────────────
elif step == 2:
    parsed = st.session_state.get("pl_parsed", {})
    year = parsed.get("year", st.session_state.get("pl_sel_year", 2026))
    month = parsed.get("month", st.session_state.get("pl_sel_month", 3))
    db_data = parsed.get("db_data", {})

    st.markdown(f"**Step 3.** {year}년 {month}월 손익실적 저장")

    # 저장할 항목 수
    total_items = len(db_data)
    factory_items = sum(1 for k in db_data if "_" in k and k.split("_")[-1] in ["gimhae", "busan", "ulsan", "gimhae2"])
    common_items = total_items - factory_items

    st.info(f"저장 대상: {year}년 {month}월, 공장별 {factory_items}개 + 공통 {common_items}개 = 총 {total_items}개 항목")

    # 주요 항목 요약 테이블
    summary_items = [
        ("매출(생산품)", "sales_prod"),
        ("매출(상품)", "sales_out"),
        ("재료비", "material"),
        ("복리후생비(변동)", "mfg_welfare"),
        ("노무비(급료)", "labor_salary"),
    ]
    summary_rows = []
    for label, prefix in summary_items:
        row = {"항목": label}
        for fcode, fname in [("gimhae","김해"), ("busan","부산"), ("ulsan","울산"), ("gimhae2","김해2")]:
            row[fname] = f"{db_data.get(f'{prefix}_{fcode}', 0):,.0f}"
        summary_rows.append(row)
    st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

    col_l, col_save, col_r = st.columns([1, 2, 1])
    with col_save:
        if st.button("DB에 저장", type="primary", use_container_width=True):
            try:
                save_monthly_pl(year, month, db_data)
                st.success(f"{year}년 {month}월 손익실적 저장 완료! (항목 {total_items}개)")
            except Exception as e:
                st.error(f"저장 실패: {e}")

    st.divider()
    render_nav_buttons(step)

# ── STEP 3: 저장 확인 ────────────────────────────────────────────────────────
elif step == 3:
    year = st.session_state.get("pl_sel_year", 2026)
    month = st.session_state.get("pl_sel_month", 3)

    st.markdown(f"**Step 4.** {year}년 {month}월 손익실적 저장 확인")
    st.caption("DB에 저장된 데이터를 조회합니다.")

    saved = load_monthly_pl(year, month)
    if not saved or not saved.get("id"):
        st.info(f"{year}년 {month}월 손익실적 데이터가 없습니다. Step 1~3에서 업로드 후 저장하세요.")
    else:
        st.success(f"{year}년 {month}월 데이터 저장 확인 (updated: {saved.get('updated_at', '?')})")

        # 공장별 매출 요약
        factories = [("gimhae", "김해"), ("busan", "부산"), ("ulsan", "울산"), ("gimhae2", "김해2")]

        rows = []
        display_fields = [
            ("판매수량(대)", "qty"),
            ("생산금액", "prod"),
            ("생산품매출", "sales_prod"),
            ("상품매출", "sales_out"),
            ("재료비", "material"),
            ("복리후생비(변동)", "mfg_welfare"),
            ("전력비", "mfg_power"),
            ("판매운반비", "selling_trans"),
            ("상품매입", "merch_purchase"),
            ("급료(노무비)", "labor_salary"),
            ("임금", "labor_wage"),
            ("상여금", "labor_bonus"),
            ("감가상각비", "fix_depr"),
            ("외주가공비", "fix_outsrc"),
        ]

        for label, prefix in display_fields:
            row = {"항목": label}
            total = 0
            for fcode, fname in factories:
                v = saved.get(f"{prefix}_{fcode}", 0) or 0
                row[fname] = v
                total += v
            row["합계"] = total
            rows.append(row)

        # 영업외 항목
        for label, field in [("영업외수익", "non_op_income"), ("영업외비용", "non_op_expense"),
                              ("이자수익", "interest_income"), ("이자비용", "interest_expense")]:
            row = {"항목": label}
            v = saved.get(field, 0) or 0
            for _, fname in factories:
                row[fname] = ""
            row["합계"] = v
            rows.append(row)

        df = pd.DataFrame(rows)
        num_cols = ["김해", "부산", "울산", "김해2", "합계"]
        for c in num_cols:
            df[c] = df[c].apply(lambda x: f"{x:,.0f}" if isinstance(x, (int, float)) else x)
        st.dataframe(df, use_container_width=True, hide_index=True)

    st.divider()
    col_l, col_c, col_r = st.columns([1, 3, 1])
    with col_l:
        if st.button("< 이전", use_container_width=True, key="prev_last"):
            st.session_state.pl_step = 2
            st.rerun()
    with col_r:
        if st.button("다음 단계: 인원·노무비 →", type="primary",
                      use_container_width=True, key="goto_next_flow"):
            try:
                st.switch_page("02_인원_노무비_입력.py")
            except Exception:
                st.info("사이드바에서 **인원·노무비** 메뉴를 클릭하세요.")
