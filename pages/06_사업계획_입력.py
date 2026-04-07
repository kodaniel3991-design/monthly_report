"""
사업계획 입력 페이지
Excel 파일 업로드 → 파싱 → 미리보기 → DB 저장
단위: 천원(KRW), 대(판매수량)
"""

import streamlit as st
import pandas as pd
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from database import save_annual_plan, load_annual_plan, load_annual_plan_as_dict
from flow_bar import render_flow_bar

st.set_page_config(page_title="사업계획 입력", layout="wide")

# ── 전체 업무 흐름 ────────────────────────────────────────────────────────
render_flow_bar(current_step=0)

# ── 사업계획 항목 정의 ─────────────────────────────────────────────────────
# (item_code, 항목명, Excel 행번호, 타입)
PLAN_ROW_MAP = [
    ("qty",             "판매수량(대)",       7,  "input"),
    ("prod",            "생산금액",           8,  "input"),
    ("sales",           "매출액",             9,  "calc"),
    ("sales_prod",      "  제품매출",         10, "input"),
    ("sales_out",       "  상품매출",         11, "input"),
    ("variable_cost",   "변동비",             12, "calc"),
    ("inv_diff",        "  제품재고증감차",    13, "input"),
    ("material",        "  재료비",           14, "input"),
    ("mfg_expense",     "  제조경비(변동)",    15, "input"),
    ("selling_trans",   "  판매운반비",       26, "input"),
    ("merch_purchase",  "  상품매입",         27, "input"),
    ("contribution",    "한계이익",           28, "calc"),
    ("fixed_cost",      "고정비",             29, "calc"),
    ("labor_cost",      "  노무비",           30, "input"),
    ("staff_cost",      "  인건비",           36, "input"),
    ("fix_mfg",         "  제조경비(고정)",    40, "input"),
    ("general_admin",   "  일반관리비",       52, "input"),
    ("operating_profit","영업이익",           60, "calc"),
    ("non_op_income",   "영업외수익",         61, "input"),
    ("non_op_expense",  "영업외비용",         63, "input"),
    ("ordinary_profit", "경상이익",           65, "calc"),
]

# 월별 컬럼 위치 (Excel 컬럼 번호, 1-based)
# 각 월 블록 = 10컬럼:
#   김해(+0) 부산(+1) RKM(+2) %(+3) 울산(+4) 김해2(+5) HKMC(+6) %(+7) 합계(+8) %(+9)
# 1~6월 시작=28, 상반기합계=88, 7~12월 시작=98
COL_OFFSETS = {
    "gimhae":  0,
    "busan":   1,
    "rkm":     2,
    "ulsan":   4,
    "gimhae2": 5,
    "hkmc":    6,
    "total":   8,
}

def _month_col(month: int, key: str) -> int:
    """월·공장/사업부 → Excel 컬럼 번호"""
    if month <= 6:
        base = 28 + (month - 1) * 10
    else:
        base = 98 + (month - 7) * 10
    return base + COL_OFFSETS[key]

# 읽을 컬럼 키 (공장 4개 + 사업부 2개 + 합계)
ALL_KEYS = [
    ("gimhae",  "김해공장"),
    ("busan",   "부산공장"),
    ("rkm",     "RKM"),
    ("ulsan",   "울산공장"),
    ("gimhae2", "김해2공장"),
    ("hkmc",    "HKMC"),
    ("total",   "합계"),
]

STEPS = [
    {"label": "Excel 업로드",   "num": "1"},
    {"label": "데이터 미리보기", "num": "2"},
    {"label": "저장",           "num": "3"},
    {"label": "연간 조회",      "num": "4"},
]


# ── 헬퍼 함수 ─────────────────────────────────────────────────────────────

def render_step_bar(current_step: int):
    """ESG 초기 설정 스타일 — 프로그레스 바 + 브레드크럼 칩"""
    total = len(STEPS)
    done = min(current_step + 1, total)
    pct = int(done / total * 100)

    # 프로그레스 카드 (컴팩트)
    st.markdown(
        f"<div style='background:white; border:0.5px solid #dddbd7; border-radius:8px; "
        f"padding:14px 20px 12px; box-shadow:0 1px 3px rgba(85,73,64,0.08); margin-bottom:12px;'>"
        # 타이틀 + 퍼센트
        f"<div style='display:flex; justify-content:space-between; align-items:center; margin-bottom:2px;'>"
        f"<div>"
        f"<span style='font-family:Sora,Pretendard,sans-serif; font-size:15px; font-weight:700; "
        f"color:#000; letter-spacing:-0.02em;'>사업계획 입력</span>"
        f"<span style='font-size:11px; color:#a8a9aa; margin-left:10px;'>Excel 업로드 → 미리보기 → 저장 → 조회</span>"
        f"</div>"
        f"<div style='font-family:Sora,Pretendard,sans-serif; font-size:20px; font-weight:800; "
        f"color:#554940; letter-spacing:-0.03em;'>{pct}%</div>"
        f"</div>"
        # 프로그레스 바
        f"<div style='height:4px; background:#eceae7; border-radius:2px; margin:8px 0 10px; overflow:hidden;'>"
        f"<div style='height:100%; width:{pct}%; background:#554940; border-radius:3px; "
        f"transition:width 0.4s cubic-bezier(0.16,1,0.3,1);'></div>"
        f"</div>"
        # 스텝 칩 (브레드크럼 — 컴팩트)
        f"<div style='display:flex; align-items:center; gap:4px; flex-wrap:wrap;'>"
        + "".join(
            # 완료
            f"<div style='display:inline-flex; align-items:center; gap:3px; padding:4px 10px; "
            f"border-radius:4px; border:1px solid #879a77; background:#f0f3ee; "
            f"font-size:11px; font-weight:500; color:#55654a;'>"
            f"&#10003; {s['num']}. {s['label']}</div>"
            f"<span style='color:#c5c6c7; font-size:10px;'>›</span>"
            if i < current_step else
            # 현재
            f"<div style='display:inline-flex; align-items:center; gap:3px; padding:4px 10px; "
            f"border-radius:4px; border:1.5px solid #554940; background:white; "
            f"font-size:11px; font-weight:700; color:#554940;'>"
            f"{s['num']}. {s['label']}</div>"
            f"<span style='color:#c5c6c7; font-size:10px;'>›</span>"
            if i == current_step else
            # 미완료
            f"<div style='display:inline-flex; align-items:center; padding:4px 10px; "
            f"border-radius:4px; border:1px solid #dddbd7; background:#f5f4f2; "
            f"font-size:11px; color:#a8a9aa;'>"
            f"{s['num']}. {s['label']}</div>"
            + (f"<span style='color:#c5c6c7; font-size:10px;'>›</span>" if i < total - 1 else "")
            for i, s in enumerate(STEPS)
        )
        + "</div>"
        f"</div>",
        unsafe_allow_html=True
    )


def render_nav_buttons(current_step: int, max_step: int = 3):
    col_l, col_c, col_r = st.columns([1, 3, 1])
    with col_l:
        if current_step > 0:
            if st.button("< 이전", use_container_width=True, key=f"prev_{current_step}"):
                st.session_state.plan_step = current_step - 1
                st.rerun()
    with col_r:
        if current_step < max_step:
            if st.button("다음 >", type="primary", use_container_width=True, key=f"next_{current_step}"):
                st.session_state.plan_step = current_step + 1
                st.rerun()


def parse_plan_excel(uploaded_file) -> dict:
    """
    사업계획 Excel 파싱
    → {month: {item_code_key: value}}
    예: {1: {"qty_gimhae": 5000, "qty_busan": 6000, "qty_rkm": 11000, ...}}
    """
    import openpyxl

    wb = openpyxl.load_workbook(uploaded_file, data_only=True)

    # "사업계획" 키워드가 포함된 시트 찾기
    target_sheet = None
    for sn in wb.sheetnames:
        if "사업계획" in sn or "계획" in sn:
            target_sheet = sn
            break
    if not target_sheet:
        target_sheet = wb.sheetnames[0]

    ws = wb[target_sheet]
    result = {"sheet_name": target_sheet, "months": {}}

    for month in range(1, 13):
        month_data = {}
        for item_code, item_name, row_num, _ in PLAN_ROW_MAP:
            for key, _ in ALL_KEYS:
                col = _month_col(month, key)
                raw = ws.cell(row_num, col).value
                val = float(raw) if raw is not None else 0.0
                month_data[f"{item_code}_{key}"] = val
        result["months"][month] = month_data

    wb.close()
    return result


def build_preview_df(parsed: dict, month: int, view: str = "월별") -> pd.DataFrame:
    """
    특정 월의 파싱 결과를 공장/사업부별 테이블로 변환
    view="월별": 공장별 상세,  view="연간": 12개월 추이 (합계 기준)
    """
    md = parsed["months"].get(month, {})

    rows = []
    for item_code, item_name, _, kind in PLAN_ROW_MAP:
        row = {"항목": item_name}
        for key, key_label in ALL_KEYS:
            v = md.get(f"{item_code}_{key}", 0) or 0
            row[key_label] = v
        rows.append(row)
    return pd.DataFrame(rows)


def build_annual_df(parsed: dict, key: str = "total") -> pd.DataFrame:
    """12개월 추이 테이블 (특정 키 기준)"""
    rows = []
    for item_code, item_name, _, kind in PLAN_ROW_MAP:
        row = {"항목": item_name}
        year_total = 0
        for m in range(1, 13):
            md = parsed["months"].get(m, {})
            v = md.get(f"{item_code}_{key}", 0) or 0
            row[f"{m}월"] = v
            year_total += v
        row["연합계"] = year_total
        rows.append(row)
    return pd.DataFrame(rows)


def format_num_df(df: pd.DataFrame) -> pd.DataFrame:
    """숫자 컬럼을 천단위 콤마 포맷"""
    out = df.copy()
    num_cols = [c for c in out.columns if c != "항목"]
    for c in num_cols:
        out[c] = out[c].apply(lambda x: f"{x:,.0f}" if x else "0")
    return out


def pct(value, base):
    return f"{value / base * 100:.1f}%" if base else "-"


# ══════════════════════════════════════════════════════════════════════════
#  메인 화면
# ══════════════════════════════════════════════════════════════════════════

# 타이틀은 render_step_bar 안에 포함됨

if "plan_step" not in st.session_state:
    st.session_state.plan_step = 0

step = st.session_state.plan_step

render_step_bar(step)

# ── STEP 0: Excel 업로드 ─────────────────────────────────────────────────
if step == 0:
    col1, col2 = st.columns([1, 4])
    with col1:
        year = st.selectbox("연도", range(2024, 2028), index=2, key="sel_year")
    st.session_state.plan_year = year

    st.info(
        "**지원 파일 형식**: `2.2 - 2026년 3월-260405.xlsx` 형태의 사업계획 파일 — "
        "월별 공장별(김해·부산·울산·김해2) + 사업부별(RKM·HKMC) 계획 데이터를 자동 추출합니다."
    )

    uploaded = st.file_uploader(
        "사업계획 Excel (.xlsx)", type=["xlsx"],
        key="plan_upload", label_visibility="collapsed"
    )

    if uploaded:
        with st.spinner("Excel 파싱 중..."):
            try:
                parsed = parse_plan_excel(uploaded)
                st.session_state.plan_parsed = parsed
                st.success(f"시트 '{parsed['sheet_name']}' 에서 12개월 × 7컬럼(공장4+사업부2+합계) 추출 완료!")

                # 간단 요약 (1월)
                m1 = parsed["months"].get(1, {})
                st.markdown(
                    f"**1월 매출 미리보기** — "
                    f"김해: {m1.get('sales_gimhae',0):,.0f} / "
                    f"부산: {m1.get('sales_busan',0):,.0f} / "
                    f"**RKM: {m1.get('sales_rkm',0):,.0f}** | "
                    f"울산: {m1.get('sales_ulsan',0):,.0f} / "
                    f"김해2: {m1.get('sales_gimhae2',0):,.0f} / "
                    f"**HKMC: {m1.get('sales_hkmc',0):,.0f}**"
                )
            except Exception as e:
                st.error(f"파싱 실패: {e}")

    # 기존 저장 데이터 표시
    existing = load_annual_plan(year)
    if existing:
        months_saved = sorted(set(r["month"] for r in existing))
        st.warning(f"{year}년 기존 데이터 있음 ({', '.join(f'{m}월' for m in months_saved)}). 업로드하면 덮어씁니다.")

    st.divider()
    col_l, col_c, col_r = st.columns([1, 3, 1])
    with col_r:
        can_next = "plan_parsed" in st.session_state and st.session_state.plan_parsed is not None
        if st.button("다음 >", type="primary", use_container_width=True,
                     disabled=not can_next, key="next_0"):
            st.session_state.plan_step = 1
            st.rerun()

# ── STEP 1: 미리보기 ─────────────────────────────────────────────────────
elif step == 1:
    parsed = st.session_state.get("plan_parsed", {})
    year = st.session_state.get("plan_year", 2026)

    st.markdown(f"**Step 2.** {year}년 사업계획 미리보기")
    st.caption(f"시트: {parsed.get('sheet_name', '?')} | 단위: 천원")

    view_mode = st.radio("조회 방식", ["월별 상세 (공장별)", "연간 추이 (12개월)"],
                          horizontal=True, key="view_mode")

    if view_mode.startswith("월별"):
        # ── 월별: 공장별 상세 테이블 ──
        preview_month = st.selectbox("월 선택", range(1, 13), index=0, key="preview_m")
        df = build_preview_df(parsed, preview_month)
        st.dataframe(format_num_df(df), use_container_width=True, hide_index=True)

        # 해당 월 주요 지표
        md = parsed["months"].get(preview_month, {})
        st.divider()
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("RKM 매출", f"{md.get('sales_rkm',0):,.0f}")
        m2.metric("HKMC 매출", f"{md.get('sales_hkmc',0):,.0f}")
        m3.metric("합계 매출", f"{md.get('sales_total',0):,.0f}")
        m4.metric("합계 경상이익", f"{md.get('ordinary_profit_total',0):,.0f}")
    else:
        # ── 연간: 12개월 추이 ──
        annual_key = st.radio(
            "조회 단위",
            ["합계", "RKM", "HKMC", "김해공장", "부산공장", "울산공장", "김해2공장"],
            horizontal=True, key="annual_key"
        )
        key_map = {"합계":"total", "RKM":"rkm", "HKMC":"hkmc",
                    "김해공장":"gimhae", "부산공장":"busan", "울산공장":"ulsan", "김해2공장":"gimhae2"}
        df = build_annual_df(parsed, key_map[annual_key])
        st.dataframe(format_num_df(df), use_container_width=True, hide_index=True)

        # 연간 합계 주요 지표
        st.divider()
        total_df = build_annual_df(parsed, "total")
        def _annual_val(item_name):
            r = total_df[total_df["항목"] == item_name]
            return r["연합계"].values[0] if len(r) else 0
        annual_sales = _annual_val("매출액")
        annual_op = _annual_val("영업이익")
        annual_ord = _annual_val("경상이익")
        m1, m2, m3 = st.columns(3)
        m1.metric("연간 매출액", f"{annual_sales:,.0f}")
        m2.metric("연간 영업이익", f"{annual_op:,.0f}", pct(annual_op, annual_sales))
        m3.metric("연간 경상이익", f"{annual_ord:,.0f}", pct(annual_ord, annual_sales))

    st.divider()
    render_nav_buttons(step)

# ── STEP 2: 저장 ─────────────────────────────────────────────────────────
elif step == 2:
    parsed = st.session_state.get("plan_parsed", {})
    year = st.session_state.get("plan_year", 2026)

    st.markdown(f"**Step 3.** {year}년 사업계획 저장")

    # item_code → item_name 매핑
    code_to_name = {}
    for item_code, item_name, _, _ in PLAN_ROW_MAP:
        for key, key_label in ALL_KEYS:
            code_to_name[f"{item_code}_{key}"] = f"{item_name.strip()}_{key_label}"

    # 저장할 데이터 요약
    total_items = 0
    for m in range(1, 13):
        total_items += len(parsed.get("months", {}).get(m, {}))

    st.info(f"저장 대상: {year}년 1~12월, 총 {total_items}개 항목")

    col_l, col_save, col_r = st.columns([1, 2, 1])
    with col_save:
        if st.button("DB에 저장", type="primary", use_container_width=True):
            try:
                saved_count = 0
                for m in range(1, 13):
                    month_data = parsed.get("months", {}).get(m, {})
                    if not month_data:
                        continue
                    items_to_save = []
                    for item_code, value in month_data.items():
                        item_name = code_to_name.get(item_code, item_code)
                        items_to_save.append({
                            "item_code": item_code,
                            "item_name": item_name,
                            "value": value or 0,
                        })
                    save_annual_plan(year, m, items_to_save)
                    saved_count += len(items_to_save)

                st.success(f"{year}년 12개월 사업계획 저장 완료! (항목 {saved_count}개)")
                pass
            except Exception as e:
                st.error(f"저장 실패: {e}")

    st.divider()
    render_nav_buttons(step)

# ── STEP 3: 연간 조회 ────────────────────────────────────────────────────
elif step == 3:
    year = st.session_state.get("plan_year", 2026)

    st.markdown(f"**Step 4.** {year}년 연간 사업계획 조회")
    st.caption("DB에 저장된 데이터를 조회합니다.")

    all_data = load_annual_plan(year)
    if not all_data:
        st.info(f"{year}년 사업계획 데이터가 없습니다. Step 1~3에서 업로드 후 저장하세요.")
    else:
        monthly = {}
        for row in all_data:
            m = row["month"]
            if m not in monthly:
                monthly[m] = {}
            monthly[m][row["item_code"]] = row["value"]

        entered_months = sorted(monthly.keys())
        st.success(f"저장 완료 월: {', '.join(f'{m}월' for m in entered_months)}")

        view_key = st.radio(
            "조회 단위",
            ["합계", "RKM", "HKMC", "김해공장", "부산공장", "울산공장", "김해2공장"],
            horizontal=True, key="view_div"
        )
        key_map = {"합계":"total", "RKM":"rkm", "HKMC":"hkmc",
                    "김해공장":"gimhae", "부산공장":"busan", "울산공장":"ulsan", "김해2공장":"gimhae2"}
        sel_key = key_map[view_key]

        rows = []
        for item_code, item_name, _, kind in PLAN_ROW_MAP:
            row = {"항목": item_name}
            year_total = 0
            for m in range(1, 13):
                v = monthly.get(m, {}).get(f"{item_code}_{sel_key}", 0) or 0
                row[f"{m}월"] = v
                year_total += v
            row["연합계"] = year_total
            rows.append(row)

        df = pd.DataFrame(rows)
        st.dataframe(format_num_df(df), use_container_width=True, hide_index=True)

    st.divider()
    col_l, col_c, col_r = st.columns([1, 3, 1])
    with col_l:
        if st.button("< 이전", use_container_width=True, key="prev_last"):
            st.session_state.plan_step = 2
            st.rerun()
    with col_r:
        if st.button("다음 단계: ① 업계동향 →", type="primary",
                      use_container_width=True, key="goto_next_flow"):
            try:
                st.switch_page("03_업계동향_입력.py")
            except Exception:
                st.info("사이드바에서 **① 업계동향** 메뉴를 클릭하세요.")
