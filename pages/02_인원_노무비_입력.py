"""
인원·노무비 입력 페이지
Excel 파일 업로드 → 파싱 → 미리보기 → DB 저장
원본: 2.1 노동생산성.xlsx (노무비(인원,근무시간) + 노무비(25-05) 시트)
단위: 천원(KRW), 명, 시간(h)
"""

import streamlit as st
import pandas as pd
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from database import save_monthly_labor, load_monthly_labor
from flow_bar import render_flow_bar

st.set_page_config(page_title="인원·노무비 입력", layout="wide")

render_flow_bar(current_step=3)

# ── Excel 셀 위치 정의 ───────────────────────────────────────────────────────
# 시트: 노무비(인원,근무시간) — sheetnames[3]
# 인원: R4(관리직), R5(생산직) × C3(RKM), C4(HKMC)
# 근무시간: R16(RKM), R17(HKMC) × C3(실작업시간), C6(잔업), C7(기본근로)
#
# 시트: 노무비(25-05) — sheetnames[2]
# 상여금: R7(사무직), R8(생산직) × C5(RKM), C10(HKMC)
# 퇴직급여: R13(사무직), R14(생산직) × C5(RKM), C10(HKMC)

STEPS = [
    {"label": "Excel 업로드",   "num": "1"},
    {"label": "데이터 미리보기", "num": "2"},
    {"label": "저장",           "num": "3"},
    {"label": "저장 확인",      "num": "4"},
]


# ── 헬퍼 함수 ────────────────────────────────────────────────────────────────

def render_step_bar(current_step: int):
    total = len(STEPS)
    done = min(current_step + 1, total)
    pct = int(done / total * 100)

    st.markdown(
        f"<div style='background:white; border:0.5px solid #dddbd7; border-radius:8px; "
        f"padding:14px 20px 12px; box-shadow:0 1px 3px rgba(85,73,64,0.08); margin-bottom:12px;'>"
        f"<div style='display:flex; justify-content:space-between; align-items:center; margin-bottom:2px;'>"
        f"<div>"
        f"<span style='font-family:Sora,Pretendard,sans-serif; font-size:15px; font-weight:700; "
        f"color:#000; letter-spacing:-0.02em;'>인원·노무비 입력</span>"
        f"<span style='font-size:11px; color:#a8a9aa; margin-left:10px;'>Excel 업로드 → 미리보기 → 저장 → 확인</span>"
        f"</div>"
        f"<div style='font-family:Sora,Pretendard,sans-serif; font-size:20px; font-weight:800; "
        f"color:#554940; letter-spacing:-0.03em;'>{pct}%</div>"
        f"</div>"
        f"<div style='height:4px; background:#eceae7; border-radius:2px; margin:8px 0 10px; overflow:hidden;'>"
        f"<div style='height:100%; width:{pct}%; background:#554940; border-radius:3px; "
        f"transition:width 0.4s cubic-bezier(0.16,1,0.3,1);'></div>"
        f"</div>"
        f"<div style='display:flex; align-items:center; gap:4px; flex-wrap:wrap;'>"
        + "".join(
            f"<div style='display:inline-flex; align-items:center; gap:3px; padding:4px 10px; "
            f"border-radius:4px; border:1px solid #879a77; background:#f0f3ee; "
            f"font-size:11px; font-weight:500; color:#55654a;'>"
            f"&#10003; {s['num']}. {s['label']}</div>"
            f"<span style='color:#c5c6c7; font-size:10px;'>›</span>"
            if i < current_step else
            f"<div style='display:inline-flex; align-items:center; gap:3px; padding:4px 10px; "
            f"border-radius:4px; border:1.5px solid #554940; background:white; "
            f"font-size:11px; font-weight:700; color:#554940;'>"
            f"{s['num']}. {s['label']}</div>"
            f"<span style='color:#c5c6c7; font-size:10px;'>›</span>"
            if i == current_step else
            f"<div style='display:inline-flex; align-items:center; padding:4px 10px; "
            f"border-radius:4px; border:1px solid #dddbd7; background:#f5f4f2; "
            f"font-size:11px; color:#a8a9aa;'>"
            f"{s['num']}. {s['label']}</div>"
            + (f"<span style='color:#c5c6c7; font-size:10px;'>›</span>" if i < total - 1 else "")
            for i, s in enumerate(STEPS)
        )
        + "</div>"
        f"</div>",
        unsafe_allow_html=True
    )


def render_nav_buttons(current_step: int, max_step: int = 3):
    col_l, col_c, col_r = st.columns([1, 3, 1])
    with col_l:
        if current_step > 0:
            if st.button("< 이전", use_container_width=True, key=f"prev_{current_step}"):
                st.session_state.labor_step = current_step - 1
                st.rerun()
    with col_r:
        if current_step < max_step:
            if st.button("다음 >", type="primary", use_container_width=True, key=f"next_{current_step}"):
                st.session_state.labor_step = current_step + 1
                st.rerun()


def parse_labor_excel(uploaded_file) -> dict:
    """
    노무비 Excel 파싱 (시트 2개 자동 인식)
    시트A: 노무비(인원,근무시간) — R3 C3="RKM" 패턴으로 감지
    시트B: 노무비(상여·퇴직) — R4 C4에 "RKM" 포함 패턴으로 감지
    """
    import openpyxl

    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    sheets = wb.sheetnames

    if len(sheets) < 2:
        raise ValueError(f"시트가 2개 이상이어야 합니다 (현재 {len(sheets)}개)")

    def cell_val(ws, row, col):
        v = ws.cell(row, col).value
        if v is None:
            return 0.0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0

    def cell_str(ws, row, col):
        v = ws.cell(row, col).value
        return str(v).strip() if v else ""

    # ── 시트 자동 인식 ──
    ws_labor = None   # 인원,근무시간
    ws_bonus = None   # 상여·퇴직
    name_labor = ""
    name_bonus = ""

    for sn in sheets:
        ws = wb[sn]
        # 인원,근무시간 시트: R3 C3 = "RKM", R4 C3 = 소수(인원수)
        if cell_str(ws, 3, 3) == "RKM" and cell_val(ws, 4, 3) > 0:
            ws_labor = ws
            name_labor = sn
        # 상여·퇴직 시트: R4 C4에 "RKM" 포함, R8 C5에 상여금 숫자
        elif "RKM" in cell_str(ws, 4, 4) and cell_val(ws, 8, 5) > 0:
            ws_bonus = ws
            name_bonus = sn

    if not ws_labor:
        raise ValueError("'노무비(인원,근무시간)' 시트를 찾을 수 없습니다. R3 C3에 'RKM' 헤더가 있어야 합니다.")
    if not ws_bonus:
        raise ValueError("'노무비(상여·퇴직)' 시트를 찾을 수 없습니다. R4 C4에 '<RKM>' 헤더가 있어야 합니다.")

    # ── 인원·근무시간 파싱 ──
    mgmt_rkm = cell_val(ws_labor, 4, 3)
    mgmt_hkmc = cell_val(ws_labor, 4, 4)
    prod_rkm = cell_val(ws_labor, 5, 3)
    prod_hkmc = cell_val(ws_labor, 5, 4)

    work_hours_rkm = cell_val(ws_labor, 16, 3)
    work_hours_hkmc = cell_val(ws_labor, 17, 3)
    overtime_gimhae = cell_val(ws_labor, 16, 6)
    overtime_busan = cell_val(ws_labor, 17, 6)
    base_hours_gimhae = cell_val(ws_labor, 16, 7)
    base_hours_busan = cell_val(ws_labor, 17, 7)

    # ── 상여·퇴직 파싱 ──
    bonus_prod_rkm = cell_val(ws_bonus, 8, 5)
    bonus_prod_hkmc = cell_val(ws_bonus, 8, 10)

    retire_mgmt_rkm = cell_val(ws_bonus, 13, 5)
    retire_mgmt_hkmc = cell_val(ws_bonus, 13, 10)
    retire_prod_rkm = cell_val(ws_bonus, 14, 5)
    retire_prod_hkmc = cell_val(ws_bonus, 14, 10)

    wb.close()

    db_data = {
        "mgmt_rkm": mgmt_rkm,
        "mgmt_hkmc": mgmt_hkmc,
        "prod_rkm": prod_rkm,
        "prod_hkmc": prod_hkmc,
        "hire_count": 0,
        "resign_count": 0,
        "work_hours_rkm": work_hours_rkm,
        "work_hours_hkmc": work_hours_hkmc,
        "overtime_gimhae": overtime_gimhae,
        "overtime_busan": overtime_busan,
        "base_hours_gimhae": base_hours_gimhae,
        "base_hours_busan": base_hours_busan,
        "bonus_prod_rkm": bonus_prod_rkm,
        "bonus_prod_hkmc": bonus_prod_hkmc,
        "retire_mgmt_rkm": retire_mgmt_rkm,
        "retire_mgmt_hkmc": retire_mgmt_hkmc,
        "retire_prod_rkm": retire_prod_rkm,
        "retire_prod_hkmc": retire_prod_hkmc,
    }

    return {
        "sheets_used": [name_labor, name_bonus],
        "db_data": db_data,
    }


def build_preview_tables(data: dict) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """미리보기 테이블 3개: 인원, 근무시간, 상여·퇴직"""
    d = data["db_data"]

    # 인원 테이블
    emp_rows = [
        {"구분": "관리직", "RKM": d["mgmt_rkm"], "HKMC": d["mgmt_hkmc"],
         "합계": d["mgmt_rkm"] + d["mgmt_hkmc"]},
        {"구분": "생산직", "RKM": d["prod_rkm"], "HKMC": d["prod_hkmc"],
         "합계": d["prod_rkm"] + d["prod_hkmc"]},
        {"구분": "합계", "RKM": d["mgmt_rkm"] + d["prod_rkm"],
         "HKMC": d["mgmt_hkmc"] + d["prod_hkmc"],
         "합계": d["mgmt_rkm"] + d["prod_rkm"] + d["mgmt_hkmc"] + d["prod_hkmc"]},
    ]

    # 근무시간 테이블
    total_hours = d["work_hours_rkm"] + d["work_hours_hkmc"]
    rkm_ratio = d["work_hours_rkm"] / total_hours * 100 if total_hours else 0
    hours_rows = [
        {"구분": "실작업시간", "RKM": f"{d['work_hours_rkm']:,.1f}h",
         "HKMC": f"{d['work_hours_hkmc']:,.1f}h",
         "합계": f"{total_hours:,.1f}h"},
        {"구분": "  잔업시간", "RKM": f"{d['overtime_gimhae']:,.1f}h (김해)",
         "HKMC": f"{d['overtime_busan']:,.1f}h (부산)",
         "합계": f"{d['overtime_gimhae'] + d['overtime_busan']:,.1f}h"},
        {"구분": "  기본근로시간", "RKM": f"{d['base_hours_gimhae']:,.1f}h (김해)",
         "HKMC": f"{d['base_hours_busan']:,.1f}h (부산)",
         "합계": f"{d['base_hours_gimhae'] + d['base_hours_busan']:,.1f}h"},
        {"구분": "근무시간 비율", "RKM": f"{rkm_ratio:.1f}%",
         "HKMC": f"{100 - rkm_ratio:.1f}%", "합계": "100%"},
    ]

    # 상여·퇴직 테이블
    bonus_rows = [
        {"구분": "상여금 (생산직)", "RKM": f"{d['bonus_prod_rkm']:,.0f}",
         "HKMC": f"{d['bonus_prod_hkmc']:,.0f}",
         "합계": f"{d['bonus_prod_rkm'] + d['bonus_prod_hkmc']:,.0f}"},
        {"구분": "퇴직급여 (사무직)", "RKM": f"{d['retire_mgmt_rkm']:,.0f}",
         "HKMC": f"{d['retire_mgmt_hkmc']:,.0f}",
         "합계": f"{d['retire_mgmt_rkm'] + d['retire_mgmt_hkmc']:,.0f}"},
        {"구분": "퇴직급여 (생산직)", "RKM": f"{d['retire_prod_rkm']:,.0f}",
         "HKMC": f"{d['retire_prod_hkmc']:,.0f}",
         "합계": f"{d['retire_prod_rkm'] + d['retire_prod_hkmc']:,.0f}"},
        {"구분": "퇴직급여 합계", "RKM": f"{d['retire_mgmt_rkm'] + d['retire_prod_rkm']:,.0f}",
         "HKMC": f"{d['retire_mgmt_hkmc'] + d['retire_prod_hkmc']:,.0f}",
         "합계": f"{d['retire_mgmt_rkm'] + d['retire_prod_rkm'] + d['retire_mgmt_hkmc'] + d['retire_prod_hkmc']:,.0f}"},
    ]

    return pd.DataFrame(emp_rows), pd.DataFrame(hours_rows), pd.DataFrame(bonus_rows)


# ══════════════════════════════════════════════════════════════════════════════
#  메인 화면
# ══════════════════════════════════════════════════════════════════════════════

if "labor_step" not in st.session_state:
    st.session_state.labor_step = 0

step = st.session_state.labor_step

render_step_bar(step)

# ── STEP 0: Excel 업로드 ─────────────────────────────────────────────────────
if step == 0:
    col1, col2, col3 = st.columns([1, 1, 3])
    with col1:
        year = st.selectbox("연도", range(2024, 2028), index=2, key="labor_year")
    with col2:
        month = st.selectbox("월", range(1, 13), index=2, key="labor_month")

    st.session_state.labor_sel_year = year
    st.session_state.labor_sel_month = month

    st.info(
        "**지원 파일**: 노무비 Excel 파일 (시트 2개) — "
        "**노무비(인원,근무시간)** 시트와 **노무비(상여·퇴직)** 시트를 자동 인식합니다."
    )

    uploaded = st.file_uploader(
        "노동생산성 Excel (.xlsx)", type=["xlsx"],
        key="labor_upload", label_visibility="collapsed"
    )

    if uploaded:
        with st.spinner("Excel 파싱 중..."):
            try:
                parsed = parse_labor_excel(uploaded)
                st.session_state.labor_parsed = parsed
                d = parsed["db_data"]
                total_emp = d["mgmt_rkm"] + d["prod_rkm"] + d["mgmt_hkmc"] + d["prod_hkmc"]
                total_hours = d["work_hours_rkm"] + d["work_hours_hkmc"]

                st.success(
                    f"시트 '{parsed['sheets_used'][0]}', '{parsed['sheets_used'][1]}' 에서 데이터 추출 완료!"
                )
                st.markdown(
                    f"**미리보기** — 종업원: **{total_emp:.0f}명** "
                    f"(RKM {d['mgmt_rkm']+d['prod_rkm']:.1f} / HKMC {d['mgmt_hkmc']+d['prod_hkmc']:.1f}) | "
                    f"실작업시간: **{total_hours:,.1f}h** | "
                    f"상여금: **{d['bonus_prod_rkm']+d['bonus_prod_hkmc']:,.0f}** 천원"
                )
            except Exception as e:
                st.error(f"파싱 실패: {e}")

    # 기존 데이터 확인
    existing = load_monthly_labor(year, month)
    if existing and existing.get("id"):
        st.warning(f"{year}년 {month}월 기존 데이터가 있습니다. 업로드하면 덮어씁니다.")

    st.divider()
    col_l, col_c, col_r = st.columns([1, 3, 1])
    with col_r:
        can_next = "labor_parsed" in st.session_state and st.session_state.labor_parsed is not None
        if st.button("다음 >", type="primary", use_container_width=True,
                     disabled=not can_next, key="next_0"):
            st.session_state.labor_step = 1
            st.rerun()

# ── STEP 1: 미리보기 ─────────────────────────────────────────────────────────
elif step == 1:
    parsed = st.session_state.get("labor_parsed", {})
    year = st.session_state.get("labor_sel_year", 2026)
    month = st.session_state.get("labor_sel_month", 3)
    d = parsed.get("db_data", {})

    st.markdown(f"**Step 2.** {year}년 {month}월 인원·노무비 미리보기")
    st.caption("단위: 명, 시간(h), 천원(KRW)")

    df_emp, df_hours, df_bonus = build_preview_tables(parsed)

    st.markdown("#### 상시종업원수 (명)")
    st.dataframe(df_emp, use_container_width=True, hide_index=True)

    st.markdown("#### 근무시간")
    st.dataframe(df_hours, use_container_width=True, hide_index=True)

    st.markdown("#### 상여금·퇴직급여 (천원)")
    st.dataframe(df_bonus, use_container_width=True, hide_index=True)

    # 검증 지표
    st.divider()
    st.markdown("**검증 지표**")
    total_emp = d.get("mgmt_rkm", 0) + d.get("prod_rkm", 0) + d.get("mgmt_hkmc", 0) + d.get("prod_hkmc", 0)
    total_hours = d.get("work_hours_rkm", 0) + d.get("work_hours_hkmc", 0)
    rkm_emp = d.get("mgmt_rkm", 0) + d.get("prod_rkm", 0)
    hkmc_emp = d.get("mgmt_hkmc", 0) + d.get("prod_hkmc", 0)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("종업원수", f"{total_emp:.0f}명")
    m2.metric("RKM 인원", f"{rkm_emp:.1f}명")
    m3.metric("HKMC 인원", f"{hkmc_emp:.1f}명")
    m4.metric("실작업시간", f"{total_hours:,.1f}h")

    if year == 2026 and month == 3:
        st.info(
            "**CLAUDE.md 기준값** — 종업원: 102명, RKM: 55.7명, HKMC: 46.3명, "
            "실작업시간: RKM 8,772h / HKMC 6,643h / 합계 15,415.5h"
        )

    st.divider()
    render_nav_buttons(step)

# ── STEP 2: 저장 ─────────────────────────────────────────────────────────────
elif step == 2:
    parsed = st.session_state.get("labor_parsed", {})
    year = st.session_state.get("labor_sel_year", 2026)
    month = st.session_state.get("labor_sel_month", 3)
    db_data = parsed.get("db_data", {})

    st.markdown(f"**Step 3.** {year}년 {month}월 인원·노무비 저장")
    st.info(f"저장 대상: {year}년 {month}월, 총 {len(db_data)}개 항목")

    # 요약
    d = db_data
    summary = [
        {"항목": "관리직 인원", "RKM": f"{d.get('mgmt_rkm',0):.1f}", "HKMC": f"{d.get('mgmt_hkmc',0):.1f}"},
        {"항목": "생산직 인원", "RKM": f"{d.get('prod_rkm',0):.1f}", "HKMC": f"{d.get('prod_hkmc',0):.1f}"},
        {"항목": "실작업시간(h)", "RKM": f"{d.get('work_hours_rkm',0):,.1f}", "HKMC": f"{d.get('work_hours_hkmc',0):,.1f}"},
        {"항목": "상여금(천원)", "RKM": f"{d.get('bonus_prod_rkm',0):,.0f}", "HKMC": f"{d.get('bonus_prod_hkmc',0):,.0f}"},
        {"항목": "퇴직급여(천원)", "RKM": f"{d.get('retire_mgmt_rkm',0)+d.get('retire_prod_rkm',0):,.0f}",
         "HKMC": f"{d.get('retire_mgmt_hkmc',0)+d.get('retire_prod_hkmc',0):,.0f}"},
    ]
    st.dataframe(pd.DataFrame(summary), use_container_width=True, hide_index=True)

    col_l, col_save, col_r = st.columns([1, 2, 1])
    with col_save:
        if st.button("DB에 저장", type="primary", use_container_width=True):
            try:
                save_monthly_labor(year, month, db_data)
                st.success(f"{year}년 {month}월 인원·노무비 저장 완료!")
            except Exception as e:
                st.error(f"저장 실패: {e}")

    st.divider()
    render_nav_buttons(step)

# ── STEP 3: 저장 확인 ────────────────────────────────────────────────────────
elif step == 3:
    year = st.session_state.get("labor_sel_year", 2026)
    month = st.session_state.get("labor_sel_month", 3)

    st.markdown(f"**Step 4.** {year}년 {month}월 인원·노무비 저장 확인")

    saved = load_monthly_labor(year, month)
    if not saved or not saved.get("id"):
        st.info(f"{year}년 {month}월 데이터가 없습니다. Step 1~3에서 업로드 후 저장하세요.")
    else:
        st.success(f"{year}년 {month}월 데이터 확인 (updated: {saved.get('updated_at', '?')})")

        rows = [
            ("관리직 인원(명)", "mgmt_rkm", "mgmt_hkmc", False),
            ("생산직 인원(명)", "prod_rkm", "prod_hkmc", False),
            ("실작업시간(h)", "work_hours_rkm", "work_hours_hkmc", True),
            ("잔업시간(h)", "overtime_gimhae", "overtime_busan", True),
            ("기본근로시간(h)", "base_hours_gimhae", "base_hours_busan", True),
            ("상여금(천원)", "bonus_prod_rkm", "bonus_prod_hkmc", True),
            ("퇴직급여-사무직(천원)", "retire_mgmt_rkm", "retire_mgmt_hkmc", True),
            ("퇴직급여-생산직(천원)", "retire_prod_rkm", "retire_prod_hkmc", True),
        ]

        table_rows = []
        for label, key_rkm, key_hkmc, is_large in rows:
            r_val = saved.get(key_rkm, 0) or 0
            h_val = saved.get(key_hkmc, 0) or 0
            if is_large:
                table_rows.append({
                    "항목": label,
                    "RKM": f"{r_val:,.1f}",
                    "HKMC": f"{h_val:,.1f}",
                    "합계": f"{r_val + h_val:,.1f}",
                })
            else:
                table_rows.append({
                    "항목": label,
                    "RKM": f"{r_val:.1f}",
                    "HKMC": f"{h_val:.1f}",
                    "합계": f"{r_val + h_val:.1f}",
                })

        st.dataframe(pd.DataFrame(table_rows), use_container_width=True, hide_index=True)

        # 입퇴사
        st.markdown(f"입사: {saved.get('hire_count', 0)}명 / 퇴사: {saved.get('resign_count', 0)}명")

    st.divider()
    col_l, col_c, col_r = st.columns([1, 3, 1])
    with col_l:
        if st.button("< 이전", use_container_width=True, key="prev_last"):
            st.session_state.labor_step = 2
            st.rerun()
    with col_r:
        if st.button("다음 단계: 노동생산성 →", type="primary",
                      use_container_width=True, key="goto_next_flow"):
            try:
                st.switch_page("04_대시보드.py")
            except Exception:
                st.info("사이드바에서 **노동생산성** 메뉴를 클릭하세요.")
